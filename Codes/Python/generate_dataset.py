# -*- coding: utf-8 -*-
"""
Produced on Fri Sep 1 20:52:11 2023
@author: KUY3IB
"""

# Import modules
import subprocess
import os
from os.path import join, isdir, abspath, exists, basename
from logging import basicConfig, error, info
from os import listdir
from re import search
from pandas import DataFrame, read_excel, concat, isna, set_option
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import json

# Set module setting
set_option('mode.chained_assignment', None)

# Hardcoded Variables
P662_DR_PATH = r"\\bosch.com\dfsrb\DfsDE\DIV\DS\CV-SA\SIV\80_Projekte_technisch\_ENH_projects_history\Projektablage alt\D.1.15 - FPT_Euro6_Tier4_CRSN3.3\D.1.15.13 Applikation\015_Dataset_Release"
P1603_DR_PATH = r"\\bosch.com\dfsrb\DfsDE\DIV\DS\CV-SA\SIV\80_Projekte_technisch\107545_FPT_C11_C13_HD_N3-22_EUVII\015_Dataset_Release"


class DatasetDatabase:
    """
    This class makes a database from dataset release folders.
    """

    def __init__(self, dir_path):

        # Initialize variables
        self.filepath = []
        self.errorpath = []
        self.errorpath2 = []
        self.database = DataFrame()

        # Start code logging
        basicConfig(
            filename='run.log',
            filemode='w',
            format='%(name)s - %(levelname)s - %(message)s'
        )

        # Add log line about process start
        info("Directory scanning has started: %s", dir_path)

        # Iterate through folders in directory
        self.gather_filepath(dir_path)

        # Using filepaths, make an excel file
        self.build_database()

        # Save the final concatenated DataFrame to a new Excel file
        self.database.to_excel('unfiltered.xlsx', index=False)

        # Modify dataframe
        self.database = self.modify_database()

        # Save the final concatenated DataFrame to a new Excel file
        self.database.to_excel('database.xlsx', index=False)

        # Modify worksheet
        self.modify_worksheet('database.xlsx')

        # save error paths in html
        # self.generate_html_from_paths(self.errorpath, "errorpath")

        # Write the list to a JSON file
        with open("error_paths", "w") as json_file:
            json.dump(self.error_path, json_file, indent=4)

    def adjust_path(self, path):
        """Adjust path to handle long file path issue in Windows."""
        if len(path) > 260:
            if path.startswith('\\\\'):
                path = '\\\\?\\UNC\\' + path[2:]
            else:
                path = '\\\\?\\' + abspath(path)
        return path

    def gather_filepath(self, dir_path):

        foldernames = listdir(dir_path)
        info("Directory has %d folders before filtering.", len(foldernames))

        for index, foldername in enumerate(foldernames):
            print(f"{index+1}/{len(foldernames)}")
            folderpath = join(dir_path, foldername)
            folderpath = self.adjust_path(folderpath)
            if not isdir(folderpath):
                continue
            try:
                filenames = listdir(folderpath)
                for filename in filenames:
                    if search(r"MergedReport.*\.xlsx", filename):
                        filepath = join(folderpath, filename)
                        filepath = self.adjust_path(filepath)
                        self.filepath.append(filepath)
            except FileNotFoundError:
                self.errorpath.append(foldername)
                error(f"File not found exception at folder {foldername}")
            except Exception as e:
                error(f"Unknown exception at the folder: {foldername}")
                error(f"Exception reason: {str(e)}")
                error("Exception details:", exc_info=True)

        info("%d folders are skipped due to error.", len(self.errorpath))

    def build_database(self):
        # Loop through each file and append its contents to the database
        pathcount = len(self.filepath)
        for index, path in enumerate(self.filepath):
            print(f"{index+1}/{pathcount}")
            try:
                # Read the Excel file into a DataFrame
                df = read_excel(
                    path,
                    engine='openpyxl',
                    sheet_name='Check_allSSD'
                )
                # Append the temporary DataFrame to the final DataFrame
                self.database = concat([self.database, df], ignore_index=True)
            except Exception as e:
                self.errorpath2.append(path)
                error(f"Unknown exception at the file: {basename(path)}")
                error(f"Exception reason: {str(e)}")

    def modify_database(self):

        df = self.database

        # -------------------------------------------------------------------

        # Drop rows where all elements are NaN
        df.dropna(how='all', inplace=True)

        # Drop columns where all elements are NaN
        df.dropna(axis=1, how='all', inplace=True)

        # Drop duplicate rows where all column values are identical
        df.drop_duplicates(inplace=True)

        # Reset index and drop old index column
        df.reset_index(drop=True, inplace=True)

        # -------------------------------------------------------------------

        for col in df.columns:

            target = False
            if col.startswith("Unnamed"):
                for idx, val in df[col].iteritems():
                    if str(val).endswith(';'):
                        df.at[idx, 'Ref'] = val
                        target = True
                if target:
                    df.drop(columns=col, inplace=True)
                    continue

            unique_vals = df[col].dropna().unique()
            if set(unique_vals).issubset({'', ' '}):
                df.drop(columns=col, inplace=True)

        # if 'Unnamed: 10' in list(df.columns):
        #     for index, row in df.iterrows():
        #         unnamed_value = row['Unnamed: 10']
        #         if str(unnamed_value).endswith(';'):
        #             df.at[index, 'Rule_ID'] = df.at[index, 'Ref']
        #             df.at[index, 'Ref'] = unnamed_value
        #     df.drop('Unnamed: 10', axis=1, inplace=True)

        if 'Unnamed: 12' in list(df.columns):
            for index, row in df.iterrows():
                unnamed_value = row['Unnamed: 12']
                if not isna(unnamed_value):
                    df.at[index, 'CalDataInfo'] = unnamed_value
            df.drop('Unnamed: 12', axis=1, inplace=True)

        if 'SSD Formula' in list(df.columns):
            for index, row in df.iterrows():
                unnamed_value = row['SSD Formula']
                if not isna(unnamed_value):
                    df.at[index, 'Rule_ID'] = unnamed_value
            df.drop('SSD Formula', axis=1, inplace=True)

        # Reset index and drop old index column
        df.reset_index(drop=True, inplace=True)

        # -------------------------------------------------------------------

        # df.to_excel('original.xlsx', index=False)

        # Drop rows where 'Label' is NaN
        df.dropna(subset=['Label'], inplace=True)

        # Drop rows where 'Value' is NaN
        df.dropna(subset=['Value'], inplace=True)

        # Drop rows where 'Recommendation' is NaN
        df.dropna(subset=['Recommendation'], inplace=True)

        # Drop rows where 'Remark' is NaN
        df.dropna(subset=['Remark'], inplace=True)

        # Drop rows where 'Decision' is NaN
        df.dropna(subset=['Decision'], inplace=True)

        # Drop rows where 'Decision' is NaN
        df.dropna(subset=['CalDataInfo'], inplace=True)

        # -------------------------------------------------------------------

        # Convert entire DataFrame to string
        df = df.astype(str)

        # Remove leading and trailing whitespaces
        df = df.applymap(str.strip)

        # Drop rows where column 'CalDataInfo' is an empty string
        df = df[df['CalDataInfo'] != '']

        # Drop rows where column 'Value' is an empty string
        df = df[df['Value'] != 'ERROR']

        # Drop rows where 'Label' has either of two specific values
        df = df.query(
            "Label != 'Missing Compliance Rules' and Label != 'SubsetCheck'")

        # Drop rows where 'Label' has either of two specific values
        df = df.query(
            "CalDataInfo != '0' and CalDataInfo != 'notAvailable'")

        # -------------------------------------------------------------------

        # Reset index after dropping rows
        df.reset_index(drop=True, inplace=True)

        # Apply the function to all cells in the DataFrame
        df = df.applymap(self.add_single_quote)

        # Sort datafrmae by Label
        df.sort_values(by='Label', inplace=True)

        # Reset index after dropping rows
        df.reset_index(drop=True, inplace=True)

        # -------------------------------------------------------------------

        return df

    # Function to conditionally prepend a single quote
    def add_single_quote(self, value):
        if value.startswith(('=', '-', '+', '@')):
            return f"'{value}"
        else:
            return value

    def modify_worksheet(self, path):
        wb = load_workbook(path)
        ws = wb.active
        tab = Table(displayName="Table1", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        for col in ws.columns:
            max_length = 0
            col = [cell for cell in col]
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width
        wb.save('database.xlsx')

    def generate_html_from_paths(self, paths, output_filename):
        """
        Generate an HTML file with clickable links from a list of paths.

        :param paths: List of paths to be turned into clickable links.
        :param output_filename: Name of the output HTML file.
        """

        # Create and open an HTML file for writing
        with open(output_filename, "w") as f:
            # Write the basic HTML structure
            f.write(
                "<html>\n<head>\n<title>Clickable Paths</title>\n</head>\n<body>\n")

            # Convert each path into a clickable link and write to the file
            for path in paths:
                f.write(f'<a href="{path}" target="_blank">{path}</a><br/>\n')

            # Close the basic HTML structure
            f.write("</body>\n</html>")


if __name__ == "__main__":

    dir_path = input("Please enter the Dataset Release directory path: ")
    dataset_object = DatasetDatabase(dir_path)

    # Construct class object and initialize
    # dataset_object = DatasetDatabase(P662_DR_PATH)

    # Construct class object and initialize
    dataset_object = DatasetDatabase(P1603_DR_PATH)


#     folder_file = [file for file in folder_file if "MergedReport" in file]

#     if len(folder_file) == 0:
#         print("\nSkipped - Folder does not have a MergedReport file.")
#         print(f"Folder name: {folder}")
#         continue

#     if len(folder_file) == 1:
#         target_file = folder_file[0]
#         target_path = os.path.join(root_path, folder, target_file)
#         root_dictionary[folder]["File Name"] = target_file
#         root_dictionary[folder]["File Path"] = target_path
#         continue

#     target_file = ""
#     max_version = -1
#     for file in folder_file:
#         match = re.search(r'tbc(\d+)', file)
#         if match:
#             version = int(match.group(1))
#             if version > max_version:
#                 max_version = version
#                 target_file = file
#         else:
#             print("\nDebug - File does not have 'tbc' version.")
#             print(f"Folder name: {folder}")
#             print(f"File name: {file}")

#     if target_file == "":
#         print("\nSkipped - Folder does not have 'tbc' version.")
#         print(f"Folder name: {folder}")

#     target_path = os.path.join(root_path, folder, target_file)
#     root_dictionary[folder]["File Name"] = target_file
#     root_dictionary[folder]["File Path"] = target_path

# with open("Database.json", "w") as complete:
#     json.dump(root_dictionary, complete)

# # %%

# dfs = []

# for folder, content in root_dictionary.items():

#     file_path = content["File Path"]

#     try:
#         df = pd.read_excel(file_path, skiprows=1, header=None)

#         index1 = df[df[df.columns[0]].str.contains(
#             'Missing Compliance Rules')].index[0]
#         index2 = df[df[df.columns[0]].str.contains('SubsetCheck')].index[0]

#         index = index1 if index1 < index2 else index2
#         df = df.iloc[:index]

#         dfs.append(df)
#     except Exception:
#         continue

# df = pd.concat(dfs, ignore_index=True)
# df.to_excel("Dataset_Database.xlsx", index=False, startrow=1, header=False)

# # %%

# wb = op.load_workbook("Dataset_Database.xlsx")
# ws = wb.active
# ws.title = "Database"

# ws["A1"].value = "Label"
# ws["B1"].value = "Value"
# ws["C1"].value = "Recommendation"
# ws["D1"].value = "Remark"
# ws["E1"].value = "Decision"
# ws["F1"].value = "Subset"
# ws["G1"].value = "SSD Resp"
# ws["H1"].value = "Case"
# ws["I1"].value = "SSD Contact"
# ws["J1"].value = "SSD Node"
# ws["K1"].value = "Rule_ID"
# ws["L1"].value = "Ref"
# ws["M1"].value = "CalDataInfo"

# wb.save("Dataset_Database.xlsx")

# # %%

# # A B C D E F G H I J K L M must be copied
# # A B C K must be checked


# def adjust_path(path):
#     """Adjust path to handle long file path issue in Windows."""
#     if len(path) > 260:
#         if path.startswith('\\\\'):
#             path = '\\\\?\\UNC\\' + path[2:]
#         else:
#             path = '\\\\?\\' + os.path.abspath(path)
#     return path


# for path in paths:
#     try:
#         directory_path = os.path.dirname(path)
#         directory_path = adjust_path(directory_path)
#         subprocess.Popen(['explorer', directory_path])
#         input("press enter")
#     except Exception as e:
#         print()
#         print(directory_path)
#         print()
