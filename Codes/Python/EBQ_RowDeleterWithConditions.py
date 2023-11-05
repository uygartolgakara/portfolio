from datetime import datetime
startTime = datetime.now()

# ------------------------------------------------------------------------------------------

import sys, os
import numpy as np
import pandas as pd
import tkinter as tk

from tkinter import filedialog
from openpyxl import load_workbook as lw
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ------------------------------------------------------------------------------------------

print("Please fill out the required parts in the interface.")

root = tk.Tk()

def close_workbook():

    root.destroy()
    sys.exit()

root.protocol("WM_DELETE_WINDOW", close_workbook)

root.title("Workbook Interface")

for i in range(1, 18):
    root.rowconfigure(i, minsize=25)

for i in range(1,5):
    root.columnconfigure(i, minsize=150)

def choose_workbook():
    global file_path

    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        workbook_name = os.path.basename(file_path)
        label1.config(text=workbook_name)

def load_workbook():

    df = pd.read_excel(file_path, sheet_name = "Tabelle6")
    columns = ["ECU_USED_FOR", "ECU_USED_IN", "CAL_PROJECT_ORGA"]

    data = []

    for column in columns:
        unique_values = df[column].unique()
        unique_values = [x for x in unique_values if x not in (None, np.nan)]
        for value in unique_values:
            data.append(f"{column} = {value}")
    listbox.delete(0, tk.END)
    for item in data:
        listbox.insert(tk.END, item)

choose_button = tk.Button(root, text="Choose Workbook", command=choose_workbook, background="#00008B", foreground = "white")
choose_button.grid(row=1, column=1, rowspan=2, columnspan=1, sticky="nsew")

load_button = tk.Button(root, text="Load Workbook", command=load_workbook, background="purple", foreground = "white")
load_button.grid(row=1, column=2, rowspan=2, columnspan=1, sticky="nsew")

label1 = tk.Label(root, text="Click load workbook to filter optional details.")
label1.grid(row=1, column=3, rowspan=2, columnspan=2, sticky="nsew")

listbox=tk.Listbox(root,selectmode="multiple",yscrollcommand=lambda f, t: scrollbar.set(f, t))
listbox.grid(row=3, column=1, rowspan=10, columnspan=4, sticky="nsew")

scrollbar = tk.Scrollbar(root, orient="vertical", command=listbox.yview)
scrollbar.grid(row=3, column=5, rowspan=10, sticky="ns")

def checkbox_handler(var, checkbox_name, label_name):
    if var.get() == 1:
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            if checkbox_name == "checkbox2":
                global path_checkbox2
                path_checkbox2 = file_path
                label_name.config(text=os.path.basename(path_checkbox2))
            if checkbox_name == "checkbox3":
                global path_checkbox3
                path_checkbox3 = file_path
                label_name.config(text=os.path.basename(path_checkbox3))
        else:
            var.set(0)  # Uncheck the checkbox if no file was selected

checkbox1 = tk.Checkbutton(root, text="Delete entries with IS_TESTCUSTOMER = True",anchor="w")
checkbox1.grid(row=13, column=1, rowspan=1, columnspan=2, sticky="nsew")

var2 = tk.IntVar()
checkbox2 = tk.Checkbutton(root, text="Delete non-existent entries in PIDC List", \
                           variable=var2, anchor="w", \
                           command=lambda: checkbox_handler(var2, "checkbox2", label2))
checkbox2.grid(row=14, column=1, rowspan=1, columnspan=2, sticky="nsew")

var3 = tk.IntVar()
checkbox3 = tk.Checkbutton(root,text="Delete non-existent entries in Attributes List", \
                           variable=var3, anchor="w", \
                           command=lambda: checkbox_handler(var3, "checkbox3", label3))
checkbox3.grid(row=15, column=1, rowspan=1, columnspan=2, sticky="nsew")

label2 = tk.Label(root, text="", anchor="w")
label2.grid(row=14, column=3, rowspan=1, columnspan=2, sticky="nsew")

label3 = tk.Label(root, text="", anchor="w")
label3.grid(row=15, column=3, rowspan=1, columnspan=2, sticky="nsew")

def close_interface():
    global selected_items
    selected_items = []
    for index in listbox.curselection():
        selected_items.append(listbox.get(index))
    root.destroy()  # Close the interface

filter_button = tk.Button(root, text="Start Process (estimated time: max. 10-15 minutes for large files)", background="green", foreground = "white", command = close_interface)
filter_button.grid(row=16, column=1, rowspan=2, columnspan=4, sticky="nsew")

root.mainloop()

# -------------------------------------------------------------------------------------------

print("\nProcess has been started. Please wait...\n")

# ------------------------------------------------------------------------------------------

con1 = False
con2 = False

# ------------------------------------------------------------------------------------------

if "file_path" in globals():
    main_path = file_path
else:
    sys.exit("Main file path is not submitted. Exiting the tool..")

# ------------------------------------------------------------------------------------------

if "path_checkbox2" in globals():
    con1 = True; pidc_path = path_checkbox2
if "path_checkbox3" in globals():
    con2 = True; attr_path = path_checkbox3

# ------------------------------------------------------------------------------------------

ws_name = ["Tabelle6","Tabelle7","Tabelle8","Tabelle9"]

dfs = pd.read_excel(main_path, sheet_name = ws_name,\
      engine='openpyxl', dtype=str, na_filter= False)

# ------------------------------------------------------------------------------------------

df6 = dfs['Tabelle6']
df7 = dfs['Tabelle7']
df8 = dfs['Tabelle8']
df9 = dfs['Tabelle9']

del dfs

# ------------------------------------------------------------------------------------------

if con1:
    pidc_df = pd.read_excel(pidc_path,sheet_name=0,engine='openpyxl', dtype='str')
    pidc_list = pidc_df['PIDC_VERS_ID'].tolist()

if con2:
    attr_df = pd.read_excel(attr_path, sheet_name = 0, engine='openpyxl', dtype='str')
    attr_list = attr_df['Attribute'].tolist()

selected = [element.split(" = ")[1] for element in selected_items]

# ------------------------------------------------------------------------------------------

print("Step 1/6 has been completed. Total runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# -------------------------------------------------------------------------------------------

# table 6 dataframe filter

if con1: df6 = df6[df6['PIDC_VERS_ID'].isin(pidc_list)]

df6 = df6.loc[df6['ACTIVE_VERSION'] == 'Y']
df6 = df6.loc[df6['IS_TESTCUSTOMER'] == 'N']
df6 = df6.loc[df6['DELETED_FLAG'] == 'N']

df6 = df6[~df6['ECU_USED_FOR'].isin(selected)]
df6 = df6[~df6['ECU_USED_IN'].isin(selected)]
df6 = df6[~df6['CAL_PROJECT_ORGA'].isin(selected)]

tab6_list = df6['PIDC_VERS_ID'].tolist()

# ------------------------------------------------------------------------------------------

print("Step 2/6 has been completed. Total runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# -------------------------------------------------------------------------------------------

# table 7 dataframe filter

if con1: df7 = df7[df7['PIDC_VERS_ID'].isin(pidc_list)]
df7 = df7[df7['PIDC_VERS_ID'].isin(tab6_list)]
df7 = df7.loc[df7['ACTIVE_VERSION'] == 'Y']
df7 = df7.loc[df7['IS_TESTCUSTOMER'] == 'N']
df7 = df7.loc[df7['IS_VARIANT'] == 'N']
if con2: df7 = df7[df7['ATTR_NAME_ENG'].isin(attr_list)]

# ------------------------------------------------------------------------------------------

print("Step 3/6 has been completed. Total runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# -------------------------------------------------------------------------------------------

# table 8 dataframe filter
if con1: df8 = df8[df8['PIDC_VERS_ID'].isin(pidc_list)]
df8 = df8[df8['PIDC_VERS_ID'].isin(tab6_list)]
df8 = df8.loc[df8['IS_SUBVARIANT'] == 'N']
if con2: df8 = df8[df8['ATTR_NAME_ENG'].isin(attr_list)]
df8 = df8.loc[df8['IS_VARIANT_DELETED'] == 'N']

# ------------------------------------------------------------------------------------------

print("Step 4/6 has been completed. Total runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# -------------------------------------------------------------------------------------------

# table 9 dataframe filter

if con1: df9 = df9[df9['PIDC_VERS_ID'].isin(pidc_list)]
df9 = df9[df9['PIDC_VERS_ID'].isin(tab6_list)]
if con2: df9 = df9[df9['ATTR_NAME_ENG'].isin(attr_list)]
df9 = df9.loc[df9['IS_VARIANT_DELETED'] == 'N']

# ------------------------------------------------------------------------------------------

print("Step 5/6 has been completed. Total runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# -------------------------------------------------------------------------------------------

if os.path.exists("TemporaryOutput_DoNotOpenOrDelete.xlsx"):
    os.remove("TemporaryOutput_DoNotOpenOrDelete.xlsx")
writer = pd.ExcelWriter('TemporaryOutput_DoNotOpenOrDelete.xlsx', engine='xlsxwriter')

df6.to_excel(writer, sheet_name='Tabelle6', index=False, header=True)
df7.to_excel(writer, sheet_name='Tabelle7', index=False, header=True)
df8.to_excel(writer, sheet_name='Tabelle8', index=False, header=True)
df9.to_excel(writer, sheet_name='Tabelle9', index=False, header=True)

writer.close() # save the output workbook

# -------------------------------------------------------------------------------------------

filename = 'TemporaryOutput_DoNotOpenOrDelete.xlsx'
worksheet_names = ['Tabelle6', 'Tabelle7', 'Tabelle8', 'Tabelle9']

data = pd.read_excel(filename, sheet_name=worksheet_names, engine='openpyxl',
                     dtype=str, na_filter= False)

# Write the data to the Excel file and autofit the columns
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    for sheet_name, df in data.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        worksheet = writer.sheets[sheet_name]

        for column in range(df.shape[1]):

            # column_letter = get_column_letter(column+1)
            # header_cell = worksheet.cell(row=1, column=column+1)
            # header_width = worksheet.column_dimensions[column_letter].width
            # max_width = max(df.iloc[:, column].astype(str).map(len).max(),\
            #                 len(header_cell.value))

            # if max_width == 0:
            #     worksheet.column_dimensions[column_letter].width = header_width * 1.3
            # else:
            #     worksheet.column_dimensions[column_letter].width = max(max_width + 2,\
            #                                                            header_width * 1.3)

            column_letter = get_column_letter(column+1)

            if df.iloc[1:, column].empty:
                max_width = 0
                worksheet.column_dimensions[column_letter].width = 8.43
            else:
                max_width = max([len(str(cell_value)) for cell_value in df.iloc[:, column]])
                worksheet.column_dimensions[column_letter].width = max_width+2

# Add a table to each worksheet
workbook = lw(filename)
i = 6
for sheet_name in worksheet_names:
    worksheet = workbook[sheet_name]

    # if not data[sheet_name].empty: # added later

    table = Table(displayName=f"Table{i}", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)

    i+=1

if os.path.exists("Output.xlsx"): os.remove("Output.xlsx")
os.rename('TemporaryOutput_DoNotOpenOrDelete.xlsx', 'Output.xlsx')

# Save the modified Excel file n876
workbook.save('Output.xlsx')

# -------------------------------------------------------------------------------------------

print("\nProcess has been completed. Program runtime: " + \
str(datetime.now()-startTime).split(".")[0]+" h/m/s\n")
