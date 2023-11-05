"""Search CDR labels inside PDF and make a DCM file from found values."""

# %% Import Modules

from pandas import options, concat, isna, DataFrame
from openpyxl import load_workbook, worksheet
from re import search, findall, sub, S
from tkinter import Tk, filedialog
from os.path import exists, isdir
from PyPDF2 import PdfFileReader
from os import remove, makedirs
from tabula import read_pdf
from shutil import rmtree
from time import strftime
from json import dump
from numpy import nan

# Set pandas to suppress notifications
options.mode.chained_assignment = None

#%% Hardcoded Variables 1

# CDR rows to hide / skip

HIDE = [
    'bassvrappl',
    'com',
    'data',
    'ddrc',
    'dfc',
    'dfes',
    'dinh',
    'diumpr',
    'dsched',
    'DSMEnv',
    'DSMRdy',
    'dtr',
    'i15031',
    'mon',
    'signals'
]

HIDE = [[item, item.lower(), item.capitalize(), item.title(), item.upper()]
        for item in HIDE]
HIDE = [item + "_" for item_list in HIDE for item in item_list]

#%% Hardcoded Variables 2

# CDR rows to show / use

SHOW = [
    'to be checked',
    'to be checked -> special case',
    'to be checked -> new alternative',
    'to be checked -> transfer rule or initial',
    'initial calibration'
]

SHOW = [[item, item.lower(), item.capitalize(), item.title(), item.upper()]
        for item in SHOW]
SHOW = [item for item_list in SHOW for item in item_list]

#%% Hardcoded Variables 3

# Headers to look for in PDF

HEADER1 = [
    'labelname',
    'label name'
]

HEADER2 = [
    'description',
    'beschreibung'
]

HEADER = [[[header1, header2],
          [header1.lower(), header2.lower()],
          [header1.capitalize(), header2.capitalize()],
          [header1.title(), header2.title()],
          [header1.upper(), header2.upper()]]
          for header2 in HEADER2 for header1 in HEADER1]

HEADER = [header_list for header_set in HEADER for header_list in header_set]

#%% Hardcoded Variables 4

# Label value identifiers in pdf

IDENTIFIER = [
    'StartWert',
    'StartValue',
    'Start value',
    'start Value',
    'StandardWert',
    'Standard value',
    'standard Value',
    'Starting value',
    'starting Value',
    r'Standardwert \| startwert',
    r'standardwert \| Startwert',
    r'StandardWert \| startWert',
    r'standardWert \| StartWert',
    r'Standardbedatung \| startwert',
    r'standardbedatung \| Startwert',
    r'StandardBedatung \| startWert',
    r'standardBedatung \| StartWert',
    r'Start Value \| standard value',
    r'start value \| Standard Value',
    r'Start value \| Standard value',
    r'Standard value \| start value',
    r'standard value \| Start value',
    r'Standard value \| Start value',
    'Default value',
    'default Value',
    'defaultValue',
    'Default wert',
    'default Wert',
    'defaultWert',
    'default',
    'Calibration start value'
]

IDENTIFIER = [[item, item.lower(), item.capitalize(), item.title(),
               item.upper()] for item in IDENTIFIER]
IDENTIFIER = [item for item_list in IDENTIFIER for item in item_list]
IDENTIFIER = "|".join(IDENTIFIER)

# Pattern to be used in dataframe label value search

PREPATTERN = "(?:" + IDENTIFIER + ")" + r"(?:\s?[:;]\s?)"

# %% Class Initialize


class DcmGenerator:
    """Use Excel and Pdf files to generate a Dcm file."""

    def __init__(self, cdr_path: str, pdf_path: str, dcm_path: str):
        """
        Use path input variables to set variables and start logging.

        Args:
            cdr_path (str): Path of xlsx type CDR file path.
            pdf_path (str): Path of pdf type SW file path.
            dcm_path (str): Path of dcm type DCM file path.

        Returns:
            None.

        """
        # Define class wide variables
        self.cdr_path = cdr_path
        self.pdf_path = pdf_path
        self.dcm_path = dcm_path

        # Initialize logging text
        self.text = 'Log file recorded at: ' + strftime('%d-%m-%Y %H:%M:%S') + '\n'

        # Delete existing output folder
        if exists("Output") and isdir("Output"):
            rmtree("Output")

        # Produce new output folder
        makedirs("Output", exist_ok=True)


# %% Is ... data?

    def is_table_data(self, label: str, df: DataFrame) -> bool:
        """
        Check if a label has tabular data contained in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (pandas.DataFrame): The dataframe to be queried.

        Returns:
            bool: True if tabular label data found, False otherwise.

        """
        try:
            check1 = label in str(df.columns[0])
            check2 = str(df.iloc[0, 0]).startswith("X: ")
            check3 = str(df.iloc[1, 0]).startswith("Y: ")
            check4 = str(df.iloc[2, 0]).startswith("Y/X")
            check5 = str(df.iloc[3, 0]).startswith("Y/X")

            checklist = [check1, check2, check3, check4, check5]
            return checklist.count(True) == 4

        except Exception:
            self.text += f"\n{label} error on is_table_data"
            return False

    def is_vector_data(self, label: str, df: DataFrame) -> bool:
        """
        Check if a label has vector data contained in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (pandas.DataFrame): The dataframe to be queried.

        Returns:
            bool: True if vector label data found, False otherwise.

        """
        try:
            check1 = label in str(df.columns[0])
            check2 = str(df.iloc[0, 0]).startswith("X:")
            check3 = str(df.iloc[1, 0]).startswith("X")
            check4 = str(df.iloc[2, 0]).startswith("VAL")

            checklist = [check1, check2, check3, check4]
            return all(checklist)

        except Exception:
            self.text += f"\n{label} error on is_vector_data"
            return False

    def is_array_data(self, label: str, df: DataFrame) -> bool:
        """
        Check if a label has array data contained in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (pandas.DataFrame): The dataframe to be queried.

        Returns:
            bool: True if array label data found, False otherwise.

        """
        try:
            df = self.filter_dataframe(df)

            pattern = label + r"\s?\[\d\]\s?="
            df_str = ''.join(map(str, df.iloc[:, 1]))
            occurrence_count = len(findall(pattern, df_str))

            check1 = list(df.columns) in HEADER
            check2 = occurrence_count > 1

            checklist = [check1, check2]
            return all(checklist)

        except Exception:
            self.text += f"\n{label} error on is_array_data"
            return False

    def is_scalar_data(self, label: str, df: DataFrame) -> bool:
        """
        Check if a label has scalar data contained in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (pandas.DataFrame): The dataframe to be queried.

        Returns:
            bool: True if scalar label data found, False otherwise.

        """
        try:
            df = self.filter_dataframe(df)

            pattern = "(?:" + IDENTIFIER + ")" + r"(?:\s?[:;]\s?)"
            df_str = ''.join(map(str, df.iloc[:, 1]))
            occurrence_count = len(findall(pattern, df_str))

            check1 = list(df.columns) in HEADER
            check2 = any(label in str(s) for s in df.iloc[:, 0])
            check3 = occurrence_count >= 1

            checklist = [check1, check2, check3]
            return all(checklist)

        except Exception:
            self.text += f"\n{label} error on is_scalar_data"
            return False

# %% Get ... Data

    def get_table_data(self, label: str, df: DataFrame) -> dict:
        """
        Gather the tabular data associated with label in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (DataFrame): Source dataframe to be queried for data.

        Raises:
            Exception: In case of compability error, return None.

        Returns:
            table_data (dict): Found label data stored as dictionary.

        """
        try:
            z_label = list(df.columns)[0].split(" ")[0]
            z_unit = findall(r"\[(.*?)\]", str(df.columns[0]))[0]

            if z_label != label:
                raise Exception

            df_split = df.iloc[:, 0].str.split(r'\s(?![^\[]*\])', expand=True)
            df = concat([df_split, df.iloc[:, 1:]], axis=1)

            if df.iloc[0, 0] == "X:":
                if not isna(df.iloc[0, 3]):
                    x_label = df.iloc[0, 1]
                    x_unit = df.iloc[0, 2][1:-1]
                    x_count = df.iloc[0, 3][1:-1]
                else:
                    x_label = "NA"
                    x_unit = df.iloc[0, 1][1:-1]
                    x_count = df.iloc[0, 2][1:-1]
            else:
                raise Exception

            if df.iloc[1, 0] == "Y:":
                if not isna(df.iloc[1, 3]):
                    y_label = df.iloc[1, 1]
                    y_unit = df.iloc[1, 2][1:-1]
                    y_count = df.iloc[1, 3][1:-1]
                else:
                    y_label = "NA"
                    y_unit = df.iloc[1, 1][1:-1]
                    y_count = df.iloc[1, 2][1:-1]
            else:
                raise Exception

            df = self.modify_dataframe(df)

            if df is None:
                raise Exception

            x_value = list(df.iloc[0, 1:])
            y_value = list(df.iloc[1:, 0])

            z_value = df.iloc[1:, 1:]
            z_value = z_value.reset_index(drop=True)
            z_value.columns = range(len(z_value.columns))

            z_count = list(z_value.shape)

            table_data = {
                'x_label': x_label,
                'y_label': y_label,
                'z_label': z_label,
                'x_unit': x_unit,
                'y_unit': y_unit,
                'z_unit': z_unit,
                'x_count': x_count,
                'y_count': y_count,
                'z_count': z_count,
                'x_value': x_value,
                'y_value': y_value,
                'z_value': z_value
            }

            return table_data

        except Exception:
            self.text += f"\n{label} error on get_table_data"
            return None

    def get_vector_data(self, label: str, df: DataFrame) -> dict:
        """
        Gather the vector data associated with label in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (DataFrame): Source dataframe to be queried for data.

        Raises:
            Exception: In case of compability error, return None.

        Returns:
            vector_data (dict): Found label data stored as dictionary.

        """
        try:
            z_label = list(df.columns)[0].split(" ")[0]
            z_unit = findall(r"\[(.*?)\]", str(df.columns[0]))[0]

            if z_label != label:
                raise Exception

            df_split = df.iloc[:, 0].str.split(r'\s(?![^\[]*\])', expand=True)
            df = concat([df_split, df.iloc[:, 1:]], axis=1)

            if df.iloc[0, 0] == "X:":
                if not isna(df.iloc[0, 3]):
                    x_label = df.iloc[0, 1]
                    x_unit = df.iloc[0, 2][1:-1]
                    x_count = df.iloc[0, 3][1:-1]
                else:
                    x_label = "NA"
                    x_unit = df.iloc[0, 1][1:-1]
                    x_count = df.iloc[0, 2][1:-1]
            else:
                raise Exception

            df = self.modify_dataframe(df, 1)

            if df is None:
                raise Exception

            x_value = list(df.iloc[0, 1:])
            x_value = [str(i) for i in x_value]
            z_value = list(df.iloc[1, 1:])
            z_value = [str(i) for i in z_value]

            z_count = len(z_value)

            if int(x_count) != z_count:
                self.text += f"\n{label} non equal x_count and z_count"

            x_count = len(x_value)

            vector_data = {
                'x_label': x_label,
                'z_label': z_label,
                'x_unit': x_unit,
                'z_unit': z_unit,
                'x_count': x_count,
                'z_count': z_count,
                'x_value': x_value,
                'z_value': z_value
            }

            return vector_data

        except Exception:
            self.text += f"\n{label} error on get_vector_data"
            return None

    def get_array_data(self, label: str, df: DataFrame) -> dict:
        """
        Gather the array data associated with label in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (DataFrame): Source dataframe to be queried for data.

        Raises:
            Exception: In case of compability error, return None.

        Returns:
            array_data (dict): Found label data stored as dictionary.

        """
        try:
            df = self.filter_dataframe(df)

            pattern = label + r"\s?\[\d\]\s?=\s?([^\s]*?)\s?\[(.*?)\]"
            combined_string = ' '.join(df.iloc[:, 1].astype(str))
            output = findall(pattern, combined_string)

            if not all(unit == output[0][1] for _, unit in output):
                raise Exception

            z_unit = output[0][1]
            z_count = len(output)
            z_value = [value for value, _ in output]

            array_data = {
                'z_unit': z_unit,
                'z_count': z_count,
                'z_value': z_value
            }

            return array_data

        except Exception:
            self.text += f"\n{label} error on get_array_data"
            return None

    def get_scalar_data(self, label: str, df: DataFrame) -> dict:
        """
        Gather scalar data associated with label in dataframe.

        Args:
            label (str): The identifier to check within the dataframe.
            df (DataFrame): Source dataframe to be queried for data.

        Raises:
            Exception: In case of compability error, return None.

        Returns:
            scalar_data (dict): Found label data stored as dictionary.

        """
        try:
            df = self.filter_dataframe(df)

            df.iloc[:, 0] = df.iloc[:, 0].str.replace('−\r', '')
            df.iloc[:, 0] = df.iloc[:, 0].ffill()
            df = df.loc[df.iloc[:, 0] == label]
            df = df.reset_index(drop=True)

            for i in range(len(df)):
                if search(PREPATTERN, str(df.iloc[i, 1])):
                    z_text = str(df.iloc[i, 1])
            z_text = findall(PREPATTERN + "(.*)", z_text)[0]

            if z_text.endswith('.'):
                z_text = z_text[:-1]

            if "[" in z_text and "]" in z_text:

                if z_text.count('"') >= 2:
                    val_type = "type0"  # text
                elif z_text.count('"') == 0:
                    val_type = "type1"  # numeric with unit
                else:
                    raise Exception

            else:
                val_type = "type2"  # numeric without unit

            z_unit, z_value = self.extract_values(val_type, z_text)
            if z_unit is None or z_value is None:
                raise Exception

            scalar_data = {
                'z_unit': z_unit,
                'z_value': z_value
            }

            return scalar_data

        except Exception:
            self.text += f"\n{label} error on get_scalar_data"
            return None


# %% Get Source Data

    def get_cdr_data(self, cdr_path: str) -> dict:
        """
        Gather data inside excel workbook as a nested dictionary.

        Args:
            cdr_path (str): Path of xlsx format Excel file.

        Returns:
            cdr_data (dict): Nested dictionary produced with Excel data.

        """
        try:

            wb = load_workbook(cdr_path)
            ws = wb["Calibration details"]

            cdr_data = self.cdr_loop_support(ws)

            cdr_missing_count = 0

            for function, data in cdr_data.items():
                cdr_missing_count += len(data)

            json_path = r"Output/CDR Labels.json"

            if exists(json_path):
                remove(json_path)

            with open(r"Output/CDR Labels.json", "w") as json_file:
                dump(cdr_data, json_file)

            function_list = list(cdr_data.keys())

            label_list = [
                key2
                for key1, val1 in cdr_data.items()
                for key2, val2 in val1.items()
            ]

            count1 = len(function_list)
            count2 = len(label_list)

            print(f"In CDR, {count1} functions with {count2} labels are missing.")
            print("CDR data has been gathered without any problem.\n")

            return cdr_data

        except Exception:
            self.text += "\nError on get_cdr_data"
            return None

    def get_pdf_data(self, pdf_path: str) -> list:
        """
        Read pdf file with tabula and returns dataframe list.

        Args:
            pdf_path (str): Path of pdf format Pdf file.

        Returns:
            list: Dataframe list obtained from Pdf file pages.

        """
        with open(pdf_path, "rb") as file:
            pdf_reader = PdfFileReader(file)
            num_pages = pdf_reader.numPages

        print(f"There are total number of {num_pages} pages are inside PDF.")

        if num_pages <= 1000:

            try:

                print(f"Currently reading pages between page 1 and page {num_pages}.")

                pdf_content = read_pdf(
                    pdf_path,
                    pages="all",
                    java_options=["-Xms4096m","-Xmx32768m"],
                    multiple_tables=True,
                    silent=True,
                    stream=True,
                    lattice=False
                    )

                return pdf_content

            except Exception:
                self.text += "\nError on get_pdf_data page: {num_pages}"
                return None

        else:

            dataframe_list = []

            for start_page in range(0, num_pages, 1000):
                end_page = min(start_page + 1000, num_pages)

                print(f"Currently reading pages between page {start_page} and page {end_page}.")

                # timestamp = dt.now().strftime("%H:%M:%S")
                # print("\nStart Page: " + str(start_page))
                # print("End Page: " + str(end_page))
                # print("Time: " + timestamp)

                try:

                    pdf_content = read_pdf(
                        pdf_path,
                        pages=f"{start_page + 1}-{end_page}",
                        java_options=["-Xms4096m","-Xmx32768m"],
                        multiple_tables=True,
                        silent=True,
                        stream=True,
                        lattice=False
                        )

                    dataframe_list += pdf_content

                except Exception:
                    self.text += "\nError on get_pdf_data: {start_page}-{end_page} page"

            return dataframe_list

    def get_dcm_data(self, dcm_path: str) -> str:
        """
        Gather DCM text data as string variable.

        Args:
            dcm_path (str): Path of dcm formatted Dcm file.

        Returns:
            dcm_data (str): Text data of Dcm file content.

        """
        try:
            with open(dcm_path, 'r') as file:
                dcm_data = file.read()
            return dcm_data

        except Exception:
            self.text += "\nError on get_dcm_data"
            return None


# %% Modify Text

    def modify_dcm_text(self, cdr_data: dict, pdf_data: dict, dcm_data: str) -> str:
        """
        Modify reference Dcm text by changing label values.

        Args:
            cdr_data (dict): Missing values dictionary from Excel file.
            pdf_data (dict): Dataframe list gathered from Pdf file.
            dcm_data (str): Unmodified text gathered from Dcm file.

        Returns:
            str: Modified Dcm text with only updated values.

        """
        found_count = 0
        misse_count = 0
        found_list = []
        misse_list = []

        for function, function_data in cdr_data.items():
            for label, label_data in function_data.items():

                found = False

                for df in pdf_data:

                    # Pass dataframe if label is not inside
                    if label not in str(df.columns) and label not in str(df.values):
                        continue

                    # Check if label has tabular data
                    if self.is_table_data(label, df):
                        data = self.get_table_data(label, df)
                        new_data = self.disect_data(data, 0, function, label, dcm_data)

                        if new_data != dcm_data:
                            dcm_data = new_data
                            found_count += 1
                            found_list.append(label)
                            found = True
                            break
                        else:
                            self.text += "\n" + label + "is table data, but value is same"

                    # Check if label has vector data
                    if self.is_vector_data(label, df):
                        data = self.get_vector_data(label, df)
                        new_data = self.disect_data(data, 1, function, label, dcm_data)

                        if new_data != dcm_data:
                            dcm_data = new_data
                            found_count += 1
                            found_list.append(label)
                            found = True
                            break
                        else:
                            self.text += "\n" + label + "is vector data, but value is same"

                    # Check if label has array data
                    if self.is_array_data(label, df):
                        data = self.get_array_data(label, df)
                        new_data = self.disect_data(data, 2, function, label, dcm_data)

                        if new_data != dcm_data:
                            dcm_data = new_data
                            found_count += 1
                            found_list.append(label)
                            found = True
                            break
                        else:
                            self.text += "\n" + label + "is array data, but value is same"

                    # Check if label has scalar data
                    if self.is_scalar_data(label, df):
                        data = self.get_scalar_data(label, df)
                        new_data = self.disect_data(data, 3, function, label, dcm_data)

                        if new_data != dcm_data:
                            dcm_data = new_data
                            found_count += 1
                            found_list.append(label)
                            found = True
                            break
                        else:
                            self.text += "\n" + label + "is scalar data, but value is same"

                if not found:
                    misse_list.append(label)
                    misse_count += 1
                    self.text += f"\nLabel not found: {label}"

        if len(found_list) > 0:
            with open(r"Output/Found Labels.json", "w") as file:
                dump(found_list, file)

        if len(misse_list) > 0:
            with open(r"Output/Missing Labels.json", "w") as file:
                dump(misse_list, file)

        print("PDF file has been searched for label values.")
        print(f"In PDF, {found_count} / {misse_count} label values are found.\n")

        return dcm_data

#%% Support functions

    def filter_dataframe(self, df: DataFrame) -> DataFrame:
        """
        Change dataframe column settings based on hardcoded headers.

        Args:
            df (DataFrame): Dataframe (may) with extra headers.

        Returns:
            DataFrame: Dataframe without extra headers.

        """
        if 'Beschreibung' in df.columns:
            if ("Label name Beschreibung" in df.columns
                or "Labelname Beschreibung" in df.columns):
                value_index = df.columns.get_loc('Beschreibung')
                df = df.iloc[:, value_index-1:value_index+1]

        if 'Description' in df.columns:
            if ("Label name Description" in df.columns
                or "Labelname Description" in df.columns):
                value_index = df.columns.get_loc('Description')
                df = df.iloc[:, value_index-1:value_index+1]

        # Check if DataFrame has only one column
        if len(df.columns) == 1:

            # Record combined column name
            header = list(df.columns)[0]

            # Split combined column name
            header_list = header.rsplit(maxsplit=1)

            # Split the values in the column at the first white space
            df = df.iloc[:, 0].str.split(n=1, expand=True)

            # Update column names
            if len(df.columns) == 2:
                df.columns = header_list

        return df  # return new dataframe

    def extract_values(self, val_type: str, z_text: str) -> tuple[str, str]:
        """
        Gather unit and data of scalar numeric or text value.

        Args:
            val_type (str): The type of the value in format.
            z_text (str): String containing value and unit.

        Raises:
            Exception: In case of uncompability raises exception.

        Returns:
            str: Found scalar unit in pdf file.
            str: Found scalar value in pdf file.

        """
        try:

            if val_type == "type0":
                pattern = r'"([^"]*)"\s?\[(.*)\]'
                z_value, z_unit = findall(pattern, z_text)[0]
                return z_unit.strip(), z_value.strip()
            elif val_type == "type1":
                pattern = r"(.*?)\s?\[(.*?)\]"
                z_value, z_unit = findall(pattern, z_text)[0]
                return z_unit.strip(), z_value.strip()
            elif val_type == "type2":
                pattern0 = r"^[+−-]?\d+(?:\.\d+)?$"
                pattern1 = r"^([+−-]?\d+(?:\.\d+)?)\s*(\S+)?"
                if search(pattern0, z_text):
                    return "−", z_text
                elif search(pattern1, z_text):
                    z_value, z_unit = findall(pattern1, z_text)[0]
                    return z_unit, z_value
            else:
                raise Exception

        except Exception:
            self.text += "\nError on extract_values"
            return None, None

    def modify_dataframe(self, df: DataFrame, row_idx=2) -> DataFrame:
        """
        Modify the pandas dataframe read from a pdf file.

        Args:
            df (DataFrame): Unmodified dataframe with faulty cells.
            row (TYPE, optional): Starting row index. Defaults to 2.

        Returns:
            df (DataFrame): Modified dataframe with correct cells.

        """
        try:
            df = df.iloc[row_idx:, :]
            df = df.reset_index(drop=True)
            df.columns = range(len(df.columns))

            for i in range(len(df)):
                row = df.iloc[i]
                # Check if all values except one are NaN in the current row
                non_nan_count = row.apply(lambda x: not isna(x)).sum()
                if non_nan_count == 1:
                    non_nan_val = row[row.apply(lambda x: not isna(x))].values[0]
                    non_nan_col = row[row.apply(lambda x: not isna(x))].index[0]

                    if str(non_nan_val).endswith(".0"):
                        non_nan_val = str(non_nan_val)[:-2]

                    # Carry non-NaN cell content to upper cell (add to upper cell string end)
                    if i > 0:
                        df.iat[i-1, df.columns.get_loc(non_nan_col)] = \
                        str(df.iat[i-1, df.columns.get_loc(non_nan_col)]) \
                        + str(non_nan_val)

                        # Change non-NaN cell to NaN
                        df.at[i, non_nan_col] = nan

                        # Check if all cells are NaN now, and drop if so
                        if df.iloc[i].isna().all():
                            df = df.drop(i)
                    break

            # Reset DataFrame index
            df = df.reset_index(drop=True)



            for i in range(len(df)):
                for j in range(len(df.columns)):

                    if isna(df.iloc[i, j]):
                        continue

                    if str(df.iloc[i, j]).endswith("-"):
                        root = str(df.iloc[i, j])

                        if j < 6:
                            inc = 1
                            for _ in range(4):
                                str_cell = str(df.iloc[i+1, j+inc])
                                if "−" in str_cell \
                                   or "." in str_cell \
                                        or str_cell == "0":
                                    inc += 1

                        else:
                            inc = 0

                        if isna(df.iloc[i+1, j+inc]):
                            continue

                        addition = str(df.iloc[i+1, j+inc])
                        modified = root[:-1] + addition
                        df.iloc[i+1, j+inc] = modified
                        df.iloc[i, j] = None

            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)
            df = df.astype(str)
            return df

        except Exception:
            self.text += "\nError on modify_dataframe"
            return None

    def cdr_loop_support(self, ws: worksheet) -> dict:
        """
        Support function for get_cdr_data function.

        Args:
            ws (worksheet): CDR workbook's data worksheet.

        Returns:
            dict: Gathered data stored as nested dictionary.

        """
        cdr_dict = {}

        for i in range(ws.min_row + 2, ws.max_row + 1):

            a_val = str(ws[f"A{i}"].value)  # function name
            b_val = str(ws[f"B{i}"].value)  # old version
            c_val = str(ws[f"C{i}"].value)  # new version
            d_val = str(ws[f"D{i}"].value)  # label name
            h_val = str(ws[f"H{i}"].value)  # calibration state
            j_val = str(ws[f"J{i}"].value)  # cal hint

            if a_val == "" \
                or d_val == "" \
                    or a_val is None \
                        or d_val is None:
                continue

            if str(b_val).lower() != "missing" \
                or str(h_val).lower() not in SHOW \
                    or str(d_val).startswith(tuple(HIDE)):
                continue

            label_data = {
                "Function Name": a_val,
                "Old Version": b_val,
                "New Version": c_val,
                "Label Name": d_val,
                "Cal. State": h_val,
                "Cal. Hint": j_val,
            }

            exist = a_val in list(cdr_dict.keys())

            if exist:
                cdr_dict[a_val][d_val] = label_data
            else:
                cdr_dict[a_val] = {d_val: label_data}

        return cdr_dict

    def disect_data(self, data:dict, mode:int, function:str, label:str, dcm_data:str):
        """
        Disect given data to get unit and value(s) and modify dcm text with new data.

        Args:
            data (dict): Gathered dictionary which contains new unit and value(s).
            mode (int): Structure of the label data dictionary.
            function (str): The function that the label is inside.
            label (str): Label with new value and unit.
            dcm_data (str): Original dcm text with original values.

        Raises:
            Exception: In uncompatibility case, raise exception..

        Returns:
            str: Modified dcm text with new unit and value(s).

        """
        try:

            # Tabular data

            if mode == 0:

                x_label = data['x_label']
                y_label = data['y_label']
                z_label = data['z_label']
                x_unit = data['x_unit']
                y_unit = data['y_unit']
                z_unit = data['z_unit']
                x_count = data['x_count']
                y_count = data['y_count']
                z_count = data['z_count']
                x_value = data['x_value']
                y_value = data['y_value']
                z_value = data['z_value']

                pattern = (
                    "KENNFELD " + z_label + ".*?"
                    "LANGNAME " + ".*?"
                    "FUNKTION " + function + ".*?"
                    "EINHEIT_X " + ".*?"
                    "EINHEIT_Y " + ".*?"
                    "EINHEIT_W " + ".*?"
                    "ST/X " + ".*?"
                    "END"
                )

                block_pattern = "(" + pattern + ")"

                if search(pattern, dcm_data, S):
                    block = findall(block_pattern, dcm_data, S)[0]
                    placeholder1 = self.make_Xtext(x_value) + "\n"
                    placeholder2 = self.make_Ytext(y_value, z_value)
                    placeholder = placeholder1 + placeholder2
                    block2 = sub("   ST/X" + ".*?" + "END",
                                 placeholder + "\n" + "END",
                                 block, flags=S)
                    dcm_data = sub(block, block2, dcm_data)

                return dcm_data

            # Vector data

            elif mode == 1:

                x_label = data['x_label']
                z_label = data['z_label']
                x_unit = data['x_unit']
                z_unit = data['z_unit']
                x_count = data['x_count']
                z_count = data['z_count']
                x_value = data['x_value']
                z_value = data['z_value']

                pattern = (
                    "KENNLINIE " + z_label + ".*?"
                    "LANGNAME " + ".*?"
                    "FUNKTION " + function + ".*?"
                    "EINHEIT_X " + ".*?"
                    "EINHEIT_W " + ".*?"
                    "ST/X " + ".*?"
                    "WERT " + ".*?"
                    "END"
                )

                block_pattern = "(" + pattern + ")"

                if search(pattern, dcm_data, S):
                    block = findall(block_pattern, dcm_data, S)[0]
                    placeholder1 = self.make_Xtext(x_value) + "\n"
                    placeholder2 = self.make_Xtext(z_value, "WERT")
                    placeholder = placeholder1 + placeholder2
                    block2 = sub("   ST/X" + ".*?" + "END",
                                 placeholder + "\n" + "END",
                                 block, flags=S)
                    dcm_data = sub(block, block2, dcm_data)

                return dcm_data

            # Array Data

            elif mode == 2:

                z_unit = data['z_unit']
                z_count = data['z_count']
                z_value = data['z_value']

                pattern = (
                    "FESTWERTEBLOCK " + label + ".*?"
                    "LANGNAME" + ".*?"
                    "FUNKTION " + function + ".*?"
                    "EINHEIT_W" + ".*?"
                    "WERT" + ".*?"
                    "END"
                )

                block_pattern = "(" + pattern + ")"

                if search(pattern, dcm_data, S):
                    block = findall(block_pattern, dcm_data, S)[0]
                    placeholder = self.make_Xtext(z_value, "WERT")
                    block2 = sub("   WERT.*?" + "END",
                                 placeholder + "\n" + "END",
                                 block, flags=S)
                    dcm_data = sub(block, block2, dcm_data)

                return dcm_data

            # Scalar Data

            elif mode == 3:

                z_unit = data['z_unit']
                z_value = data['z_value']

                if z_unit == "−":
                    z_unit = "-"

                if z_value.replace(" ", "").isalpha():
                    regex_mode = "text"
                else:
                    regex_mode = "value"

                pattern = (
                    "FESTWERT " + label + ".*?"
                    "LANGNAME " + ".*?"
                    "FUNKTION " + function + ".*?"
                    "EINHEIT_W " + ".*?"
                    "(?:WERT|TEXT)" + ".*?"
                    "END"
                )

                block_pattern = "(" + pattern + ")"

                if search(pattern, dcm_data, S):
                    block = findall(block_pattern, dcm_data, S)[0]

                    if regex_mode == "text":
                        block2 = sub(r'TEXT ".*?"', rf'TEXT "{z_value}"', block)
                        dcm_data = sub(block, block2, dcm_data)
                    else:
                        # DISCUSSION TOPIC - BattU_uVccSRCMax_C
                        block1 = sub(r'EINHEIT_W ".*?"', rf'EINHEIT_W "{z_unit}"', block)
                        block2 = sub(r'WERT [+−-]?\d+(?:\.\d+)?', rf'WERT {z_value}', block1)
                        # dcm_data = sub(block, block1, dcm_data)
                        dcm_data = sub(block, block2, dcm_data)

                return dcm_data

            else:
                raise Exception

        except Exception:
            self.text += f"\nError on disect_data, label: {label}\n"
            return dcm_data  # return ""

    def make_Xtext(self, x_list, label="ST/X"):
        """While writing DCM text, make value string."""
        Xtext = ""

        for i, value in enumerate(x_list):

            if i % 6 == 0:
                Xtext += label

            Xtext += "   " + str(value)

            if (i+1) % 6 == 0 or i == len(x_list) - 1:
                Xtext += "\n"

        lines = Xtext.split("\n")
        lines = ["   " + line for line in lines]
        Xtext = "\n".join(lines)
        Xtext = Xtext.rstrip()

        return Xtext

    def make_Ytext(self, y_list, z_df):
        """While writing DCM text, make value string for related cases."""
        Ytext = ""

        for idx, y in enumerate(y_list):
            tip = "" if idx == 0 else "\n"
            Ytext += f"{tip}   ST/Y   {y}\n"
            Ytext += self.make_Xtext(list(z_df.iloc[idx, :]), "WERT")

        return Ytext

# %% Main


def main():
    """Start of program to produce a DCM file from pdf and CDR files."""
    file_window = Tk()
    file_window.withdraw()

    print("\nWelcome, please select required files when prompted:\n")

    print("Select CDR file in (.xlsx) format with file browser")
    cdr_path = filedialog.askopenfilename(filetypes=[("CDR File", "*.xlsx")])
    print("Select PDF file in (.pdf) format with file browser")
    pdf_path = filedialog.askopenfilename(filetypes=[("PDF File", "*.pdf")])
    print("Select DCM file in (.dcm) format with file browser")
    dcm_path = filedialog.askopenfilename(filetypes=[("DCM File", "*.dcm")])

    if cdr_path == "" or pdf_path == "" or dcm_path == "":
        raise RuntimeError("You need to give CDR, PDF and DCM files to this tool.")

    global tool
    tool = DcmGenerator(cdr_path, pdf_path, dcm_path)

    print("\nInput files have been gathered without any problem.\n")

    try:

        global cdr_data, pdf_data, dcm_data

        # Read CDR Data

        cdr_data = tool.get_cdr_data(cdr_path)
        if cdr_data is None:
            raise Exception("CDR data could not be read.")

        # print("\nCDR data has been gathered without any problem.\n")

        # Read PDF Data

        pdf_data = tool.get_pdf_data(pdf_path)
        if pdf_data is None:
            raise Exception("PDF data could not be read.")

        print("PDF data has been read without any problem.\n")

        # Modify DCM Data

        dcm_data = tool.get_dcm_data(dcm_path)
        if dcm_data is None:
            raise Exception("DCM data could not be read.")

        print("DCM data has been gathered without any problem.\n")

        # Modify DCM Text
        dcm_data = tool.modify_dcm_text(cdr_data, pdf_data, dcm_data)

        print("New DCM text has been produced without any problem.")

        # Make DCM File

        if exists(r"Output/Missing.DCM"):
            remove(r"Output/Missing.DCM")

        with open(r'Output/Missing.DCM', 'x', encoding='utf-8') as file:
            file.write(dcm_data)

        print("DCM file has been modified without any problem.\n")

        if exists(r"Output/Log.txt"):
            remove(r"Output/Log.txt")

        with open(r'Output/Log.txt', 'x', encoding='utf-8') as file:
            file.write(tool.text)

        print("Log file has been produced without any problem.")


    except Exception as e:
        raise RuntimeError("Critical error: " + str(e))


if __name__ == "__main__":
    main()
