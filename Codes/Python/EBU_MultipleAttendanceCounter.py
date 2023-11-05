# --------------------------------------------------------------------------------------------------------------------

"""
Produced on Wed May 17 13:23:30 2023
@author: KUY3IB
-*- coding: utf-8 -*-
"""

# --------------------------------------------------------------------------------------------------------------------

from datetime import datetime
startTime = datetime.now()

import re
import os
import sys
import time
import pandas as pd
import openpyxl as op
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import TableStyleInfo

# --------------------------------------------------------------------------------------------------------------------

def format_name(name):

    if "(" in name and ")" in name:

        name = name.split('(')[0].strip()

        if ',' in name:

            last_name, first_name = name.split(',')
            return f'{first_name.strip()} {last_name.strip()}'

        else:

            name_list = name.split()

            if len(name_list) == 2:
                last_name = name_list[0].strip()
                first_name = name_list[1].strip()
                return first_name + " " + last_name

            elif len(name_list) > 2:
                last_name = name_list[0].strip()
                first_name = " ".join(name_list[1:])
                return first_name + " " + last_name

            else:
                return name.strip()
    else:
        return name.strip()

# --------------------------------------------------------------------------------------------------------------------

def handle_special_cases(name):

    if re.match(r'\+?\d[\d -]{7,}\d$', name): return name
    name = re.sub(r'\b\d+\b$', '', name)
    name = re.sub(r'\d$', '', name)

    return name.strip()

# --------------------------------------------------------------------------------------------------------------------

cwd_path = os.getcwd()

output_file = 'attendance_report.xlsx'
output_path = os.path.join(cwd_path, output_file)
if os.path.exists(output_file):  os.remove(output_file)

xlsx_files = [filename for filename in os.listdir(cwd_path) if filename.endswith(".xlsx")]

if len(xlsx_files) == 0:
    print("There are no xlsx files near executable tool. Please put xlsx file(s) near executable tool!")
    time.sleep(3); sys.exit("Terminating..")

xlsx_paths = [os.path.join(cwd_path, filename) for filename in xlsx_files]

# --------------------------------------------------------------------------------------------------------------------

wb = op.Workbook()
ws = wb.active
ws.title = "Attendance"
ws["A1"].value = "Student Name"

# --------------------------------------------------------------------------------------------------------------------

for i, xlsx_path in enumerate(xlsx_paths):

    df = pd.read_excel(xlsx_path)

    df['Timestamp'] = pd.to_datetime(df['Timestamp'])

    df['Full Name'] = df['Full Name'].str.replace(r"\(Guest\)", "")
    df['Full Name'] = df['Full Name'].str.replace(r"\.", " ")
    df['Full Name'] = df['Full Name'].str.replace("FIXED-TERM ", "")
    df['Full Name'] = df['Full Name'].apply(format_name).str.title()
    df['Full Name'] = df['Full Name'].apply(handle_special_cases)

    df['User Action'] = df['User Action'].str.replace("Joined before", "Joined")
    df = df.sort_values(['Full Name', 'Timestamp'])

    ws.cell(row=1, column=i+2).value = f"Attendance {i+1} (min)"

# --------------------------------------------------------------------------------------------------------------------

    result = pd.DataFrame(columns=['Full Name', 'Attendance Duration'])

    for name in df['Full Name'].unique():

        temp_df = df[df['Full Name'] == name]
        if temp_df.shape[0] % 2 != 0: continue

        joined_timestamps = temp_df[temp_df['User Action'] == 'Joined']['Timestamp'].tolist()
        left_timestamps = temp_df[temp_df['User Action'] == 'Left']['Timestamp'].tolist()

        if len(joined_timestamps) != len(left_timestamps): continue

        durations = [(left - joined).total_seconds() for joined, left in zip(joined_timestamps, left_timestamps)]
        total_minutes = int(sum(durations) / 60)

# --------------------------------------------------------------------------------------------------------------------

        student_found = False

        for ii in range(2,ws.max_row+1):
            if ws.cell(ii, 1).value == name:
                student_found = True
                ws.cell(ii, i+2).value = total_minutes
                break

        if not student_found:
            max_row = ws.max_row
            ws.cell(max_row + 1, 1).value = name
            ws.cell(max_row + 1, i+2).value = total_minutes

# --------------------------------------------------------------------------------------------------------------------

ws.cell(1, ws.max_column+1).value = "Total Attendance (min)"

for i in range(2, ws.max_row+1):
    sum = 0
    for ii in range(2, ws.max_column):
        if ws.cell(i, ii).value is not None and ws.cell(i, ii).value != "":
            sum += int(ws.cell(i, ii).value)
    ws.cell(i, ws.max_column).value = sum

# --------------------------------------------------------------------------------------------------------------------

table_ref = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
tab = op.worksheet.table.Table(displayName="AttendanceTable", ref=table_ref)
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight12",
showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

# --------------------------------------------------------------------------------------------------------------------

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
green_fill =PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green

for i in range(2, ws.max_row + 1):
    if ws.cell(i, ws.max_column).value is not None:
        if int(ws.cell(i, ws.max_column).value) < 60 * len(xlsx_files):
            ws.cell(i, ws.max_column).fill = red_fill
        if int(ws.cell(i, ws.max_column).value) >= 60 * len(xlsx_files):
            ws.cell(i, ws.max_column).fill = green_fill

# --------------------------------------------------------------------------------------------------------------------

for i, column in enumerate(ws.columns, start=1):
    for cell in column:
        if i == 1:
            ws.column_dimensions[cell.column_letter].width = 30
        elif i == ws.max_column:
            ws.column_dimensions[cell.column_letter].width = 30
        else:
            ws.column_dimensions[cell.column_letter].width = 15

# --------------------------------------------------------------------------------------------------------------------

wb.save(output_file)
os.startfile(output_file)

# --------------------------------------------------------------------------------------------------------------------

print("Program runtime: " + str(datetime.now()-startTime))

# --------------------------------------------------------------------------------------------------------------------
