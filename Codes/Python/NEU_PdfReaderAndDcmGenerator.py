def dcm_missing(cdr_path, sw_path, ram = 2, dcm_path = ""):

    """ This function produce dcm file from a cdr file (excel) and a sw document (pdf)"""

#%%
    from datetime import datetime
    startTime = datetime.now()

    import openpyxl as op
    import tabula as tb
    import pandas as pd
    import json, re

#%%

    def is_tableData(label, df):

        if df.shape[0] < 3 or df.shape[1] < 1: return False

        check1 = True if label in str(df.columns[0]) else False
        check2 = True if str(df.iloc[0,0]).startswith("X: ") else False
        check3 = True if str(df.iloc[1,0]).startswith("Y: ") else False
        check4 = True if str(df.iloc[2,0]).startswith("Y/X") else False
        check5 = True if str(df.iloc[3,0]).startswith("Y/X") else False

        checklist = [check1, check2, check3, check4, check5]
        return True if checklist.count(True) == 4 else False

    # ---------------------------------------------------------------------------------------

    def get_tableData(df):

        label = str(df.columns[0]).split("−")[0].rstrip()
        label_unit = re.search(r"\[(.*?)\]", str(df.columns[0])).group(1)

        df_split = df.iloc[:,0].str.split(expand=True)

        for i in range(2, df_split.shape[0]):
            offset = 0; offset_found = False
            for j in range(0, df_split.shape[1]):
                if str(df_split.iloc[i,j]).endswith("-"):
                    if not offset_found:
                        for k in range(j+1, df_split.shape[1]):
                            if not str(df_split.iloc[i + 1, k]).startswith("−"):
                                offset = k-j; offset_found = True; break
                    next_col = j + offset
                    if str(df_split.iloc[i+1, next_col]) == "0": next_col += 1
                    df_split.iloc[i+1,next_col] = \
                        df_split.iloc[i,j][:-1]+df_split.iloc[i+1, next_col]
                    df_split.iloc[i, j] = None

        df = pd.concat([df_split, df.iloc[:,1:]], axis=1)

        for i in range(2, df.shape[0]-1):
            for j in range(df.shape[1]):
                if str(df.iloc[i, j]).endswith("-") and not pd.isnull(df.iloc[i, j]):
                    df.iloc[i + 1, j]=str(df.iloc[i, j]).strip("-")+str(df.iloc[i + 1, j])
                    df.iloc[i, j] = float('nan')

        df = df[df.notnull().any(axis=1)]
        df.reset_index(drop=True, inplace=True)
        df.columns = list(range(df.shape[1]))

        zero_df = pd.DataFrame([['0'] * df.shape[1]], columns=df.columns)
        df = pd.concat([df.iloc[:3], zero_df, df.iloc[3:]]).reset_index(drop=True)

        if df.iloc[0,0]=="X:":
            x_unit = re.search(r'\[(.*?)\]', str(df.iloc[0,2])).group(1)
            x_num = re.search(r'\[(.*?)\]', str(df.iloc[0,3])).group(1)

        if df.iloc[1,0]=="Y:":
            y_unit = re.search(r'\[(.*?)\]', str(df.iloc[1,2])).group(1)
            y_num = re.search(r'\[(.*?)\]', str(df.iloc[1,3])).group(1)

        return label, label_unit, x_num, x_unit, y_num, y_unit, df

    # ---------------------------------------------------------------------------------------

    def make_xText(val_list, label="ST/X"):

        text = ""

        for i, value in enumerate(val_list):

            if i % 6 == 0: text += label
            text+= "   " + str(value)
            if (i+1) % 6 == 0 or i == len(val_list) - 1: text += "\n"

        lines = text.split("\n")
        lines = ["   " + line for line in lines]
        text = "\n".join(lines)
        text = text.rstrip()

        return text

    # ---------------------------------------------------------------------------------------

    def make_yText(df):

        text = ""

        for i in range(3, df.shape[0]):
            text += "ST/Y   " + df.iloc[i,0] + "\n"
            text += make_xText(list(df.iloc[i,1:]),"WERT")

        lines = text.split("\n")
        lines = ["   " + line for line in lines]
        text = "\n".join(lines)

        return text

    # ---------------------------------------------------------------------------------------

    def is_arrayData(label,df):

        if df.shape[0] != 2 or df.shape[1] < 1: return False # satır sayısı

        check1 = True if label in str(df.columns[0]) else False
        check2 = True if str(df.iloc[0,0]) == "X" else False
        check3 = True if str(df.iloc[1,0]) == "VAL" else False

        checklist = [check1, check2, check3]
        return True if all(checklist) else False

    # ---------------------------------------------------------------------------------------

    def get_arrayData(df):

        label = str(df.columns[0]).split("−")[0].rstrip()
        bracket = re.findall(r"\[(.*?)\]", str(df.columns[0]))
        [label_unit, x_unit, x_num] = bracket

        return label, label_unit, x_unit, x_num, df

    # ---------------------------------------------------------------------------------------

    def is_listData(label, df, tab_pick):

        pat = label + r"\[\d\] = "

        check1 = True if list(df.columns) in tab_pick else False
        check2 = True if df.iloc[:, 0].str.contains(label).any() else False
        check3 = True if df.iloc[:, 1].apply(lambda x: re.search(pat, x) \
                                             is not None).any() else False
        checklist = [check1, check2, check3]
        return True if all(checklist) else False

    # ---------------------------------------------------------------------------------------

    def get_listData(label, df):
        for i in range(df.shape[0]):
            if str(df.iloc[i,0]) == label:
                data = re.findall(label + r"\[\d+\]\s=\s(.*?)\[(.*?)\]", str(df.iloc[i,1]))
                break

        if all(x[1] == data[0][1] for x in data): val_unit = data[0][1]
        val_list = [x[0] for x in data]

        return label, val_unit, val_list

    # ---------------------------------------------------------------------------------------

    def is_singleData(label, df, tab_pick, prepattern):

        pat = label + r"\[\d\] = "

        check1 = True if list(df.columns) in tab_pick else False
        check2 = True if df.iloc[:, 0].str.contains(label).any() else False
        check3 = True if df.iloc[:, 1].apply(lambda x: re.search(prepattern,x) \
                                             is not None).any() else False
        check4 = False if df.iloc[:, 1].apply(lambda x: re.search(pat, x) \
                                              is not None).any() else True

        checklist = [check1, check2, check3, check4]
        return True if all(checklist) else False

    # ---------------------------------------------------------------------------------------

    def get_singleData(label, df, prepattern):

        for i in range(df.shape[0]):
            if str(df.iloc[i,0]) == label:
                if re.search(prepattern, str(df.iloc[i,1])):
                    v_text = re.search(prepattern + "(.*)", str(df.iloc[i,1])).group(1)

                    # -----------------------------------------------------------------
                    # there are 2 square brackets in the value text (e.g. 200 [ km/h ])
                    # -----------------------------------------------------------------

                    if ("[" in v_text) and ("]" in v_text):

                        if v_text.count('"') == 2:

                            # text is a string value (e.g. "sub text"[])
                            # ******************************************

                            pattern = r'"([^"]*)"[\s]*\[(.*)\]'
                            matches = re.match(pattern, v_text)

                            value = matches.group(1)
                            unit = matches.group(2)

                            value = value.strip()
                            unit = unit.strip()

                            # return unit, value

                        else:

                            # text is a numeric value with unit (e.g. 125[km/h])
                            # **************************************************

                            value, unit = re.findall(r"(.*?)\s?\[(.*?)\]", v_text)[0]

                            value = value.strip()
                            unit = unit.strip()

                            unit = format_unit(unit)

                            # return unit, value

                    # ------------------------------------------------------------
                    # there are no square brackets in the value text (e.g. 42 rpm)
                    # ------------------------------------------------------------

                    else:

                        # text is a number without unit (e.g. -295.035)
                        # *********************************************

                        if re.search(r"^[+−-]?\d+(?:\.\d+)?$", v_text):

                            value = v_text; unit = "−"
                            value = value.strip(); unit = unit.strip()

                            unit = format_unit(unit)

                            # return unit, value


                        # text is a number with a unit (e.g. 600 rpm)
                        # *******************************************

                        match = re.match(r"^([+−-]?\d+(?:\.\d+)?)\s*(\S+)?$", v_text)

                        if match:

                            value, unit = match.groups()
                            value = value.strip(); unit = unit.strip()

                            unit = format_unit(unit)

                            # return unit, value

    # ---------------------------------------------------------------------------------------

    def format_unit(unit):

        if unit == "" : unit = "−"

        unit = unit.replace("°" ,"deg ")
        unit = unit.replace("μ" ,"u")
        unit = unit.replace("μ", "u")
        unit = unit.replace("." ,"")

        return unit

    # ---------------------------------------------------------------------------------------

    # def is_alleData(label,df):

    #     check1 = True if list(df.columns) == ["Labelname", "Beschreibung"] else False
    #     check2 = True if str(df.iloc[0,0]) == label else False
    #     check3 = True if re.search(r"\(alle\)", str(df.iloc[0,1])) else False
    #     check4 = True if re.search(r"X:", str(df.iloc[0,1])) else False
    #     check5 = True if not re.search(r"Y:", str(df.iloc[0,1])) else False

    #     checklist = [check1, check2, check3, check4, check5]
    #     return True if checklist.count(True) == 5 else False

    # def get_alleData(df):

    #     label = str(df.iloc[0,0])
    #     label_unit = re.search(r"\(alle\).*?\[(.*?)\]", str(df.iloc[0,1])).group(1).strip()
    #     label_val = ?

    #     [x_unit,x_num] = re.findall(r"X:.*?\[(.*?)\]\s?\[(.*?)\]", str(df.iloc[0,1]))[0]
    #     x_unit = x_unit.strip(); x_num = x_num.strip()

    #     return label, label_unit, x_num, x_unit, df

#%% import modules

    pd.options.mode.chained_assignment = None # to hide copy warnings

#%% define variables

    cal_show = ["to be checked"                             ,
                "to be checked -> special case"             ,
                "to be checked -> new alternative"          ,
                "to be checked -> transfer rule or initial" ]

    cal_show = [[cal, cal.lower(), cal.capitalize(), cal.upper()] for cal in cal_show]
    cal_show = [cal for cal_group in cal_show for cal in cal_group]

    # ---------------------------------------------------------------------------------------

    lab_hide = ["Bassvrappl"  ,
                "COM"         ,
                "DATA"        ,
                "DDRC"        ,
                "DFC"         ,
                "DFES"        ,
                "DINH"        ,
                "DIUMPR"      ,
                "DSCHED"      ,
                "DSMEnv"      ,
                "DSMRdy"      ,
                "DTR"         ,
                "I15031"      ,
                "Signals"     ]

    lab_hide = [lab + "_" for lab in lab_hide]
    lab_hide = [[lab, lab.lower(), lab.capitalize(), lab.upper()] for lab in lab_hide]
    lab_hide = [lab for lab_group in lab_hide for lab in lab_group]

    # ---------------------------------------------------------------------------------------

    tab_pick = [["Label name" , "Description"  ],
                ["Labelname"  , "Beschreibung" ]]

    # ---------------------------------------------------------------------------------------

    key_look = ["StartWert"                     ,
                "Start Value"                   ,
                "StandardWert"                  ,
                "Standard Value"                ,
                "Starting Value"                ,
                "StandardWert \| StartWert"     ,
                "Start Value \| Standard Value" ]

    key_look = [[key, key.lower(), key.capitalize(), key.upper()] for key in key_look]
    key_look = [key for key_group in key_look for key in key_group]
    key_look.append("Start value") # wild case
    key_look = "|".join(key_look)

    # ---------------------------------------------------------------------------------------

    prepattern = "(?:" + key_look + ")" + r"(?:\s?[:;]\s?)"

    print("\nStep 1 / 6 has been completed, total runtime: " + \
    str(datetime.now()-startTime).split(".")[0]+" h/m/s\n")

#%% gather data

    cdr_dict = {}
    cdr_missing = 0

    wb = op.load_workbook(cdr_path)
    ws = wb["Calibration details"]

    for i in range(ws.min_row + 2, ws.max_row + 1):

        a_val = str(ws[f"A{i}"].value) # function name
        b_val = str(ws[f"B{i}"].value) # old version
        c_val = str(ws[f"C{i}"].value) # new version
        d_val = str(ws[f"D{i}"].value) # label name
        h_val = str(ws[f"H{i}"].value) # calibration

        if a_val == "" or a_val == "None": continue
        if d_val == "" or d_val == "None": continue

        if str(b_val).lower() != "missing": continue
        if str(h_val).lower() not in cal_show: continue
        if str(d_val).startswith(tuple(lab_hide)): continue

        lab_data = {"Old Version": b_val,"New Version": c_val}

        inside = a_val in list(cdr_dict.keys())
        if inside: cdr_dict[a_val][d_val] = lab_data
        if not inside: cdr_dict[a_val] = {d_val: lab_data}

    for function, data in cdr_dict.items(): cdr_missing += len(data)

#%% make json

    if os.path.exists("missing_cdr.json"): os.remove("missing_cdr.json")
    with open("missing_cdr.json","w") as f: json.dump(cdr_dict,f)

    list2 = [key2 for key1, val1 in cdr_dict.items() for key2, val2 in val1.items()]
    list1 = list(cdr_dict.keys()); num1 = len(list1); num2 = len(list2)

    print("JSON output file has been generated in same directory. Information:")
    print(f"In CDR file there are {num1} functions having labels with missing values.")
    print(f"These functions have a total number of {num2} labels with missing values.")

    print("\nStep 2 / 6 has been completed, total runtime: " + \
    str(datetime.now()-startTime).split(".")[0]+" h/m/s\n")

#%% longest runtime

    """ Gather tables from pdf as dataframe list. If specified memory - RAM is not enough
        multiply the RAM value by 2. In a process it may go as 1-2-4-8-16 (Limited at 32)"""

    read = False

    while read == False and ram < 18:

        try:
            print(f"Using {ram} GB of RAM to read tables inside pdf. Please wait...")
            dfs=tb.read_pdf(sw_path, pages="all",java_options=(f"-Xmx{ram}g"),
                            silent=True, stream = True)
            read = True
        except:
            print(f"{ram} GB of RAM was not enough. Adding 2 GB to size. Please wait...")
            ram = ram + 2

        if ram == 18: sy.exit("16 GB RAM was not enough for process. Exiting the program.")

    print(f"PDF reading was successful with {ram} GB RAM. Continuing process...\n")

    print("Step 3 / 6 has been completed, total runtime: " + \
    str(datetime.now()-startTime).split(".")[0]+" h/m/s")

#%% reference dcm

    if dcm_path != "" :
        with open(dcm_path,'r') as file: dcm_data = file.read()
        compare = True
    else:
        compare = False

    print("Step 4 / 6 has been completed, total runtime: " + \
    str(datetime.now()-startTime).split(".")[0]+" h/m/s")

#%% write dcm

    text = "" ; dcm_dict = {}

    extra1 = ['Unnamed: 0','Labelname','Beschreibung']
    extra2 = ['Unnamed: 0','Unnamed: 1','Labelname','Beschreibung']

    for function,function_data in cdr_dict.items():
        for label,label_data in function_data.items():
            for df in dfs: # loop through all dataframes

                if not re.search(label, str(df.values)): continue

                if list(df.columns) == extra1: df = df.iloc[:,-2:]
                if list(df.columns) == extra2: df = df.iloc[:,-2:]

                df.iloc[:, 0] = df.iloc[:, 0].str.replace('−\r', '')
                df.iloc[:,0] = df.iloc[:,0].ffill()

                if is_tableData(label, df):

                    label, l_unit, x_num, x_unit, y_num, y_unit, df = get_tableData(df)

                    if compare:
                        match1=re.search(f'{label}.*?EINHEIT_X\s"(.*?)"', dcm_data, re.S)
                        match2=re.search(f'{label}.*?EINHEIT_Y\s"(.*?)"', dcm_data, re.S)
                        match3=re.search(f'{label}.*?EINHEIT_W\s"(.*?)"', dcm_data, re.S)

                        if match1:
                            if x_unit != match1.group(1): x_unit = match1.group(1)
                        if match2:
                            if y_unit != match2.group(1): y_unit = match2.group(1)
                        if match3:
                            if l_unit != match3.group(1): l_unit = match3.group(1)

                    text += f"KENNFELD {label} {x_num} {y_num}\n"       + \
                            f"   FUNKTION {function}\n"                 + \
                            f'   EINHEIT_X "{x_unit}"\n'                + \
                            f'   EINHEIT_Y "{y_unit}"\n'                + \
                            f'   EINHEIT_W "{l_unit}"\n'                + \
                            make_xText(list(df.iloc[2,1:]))+ "\n"       + \
                            make_yText(df) + "\n"                       + \
                            "END\n\n"

                if is_arrayData(label,df):

                    label, l_unit, x_unit, x_num, df = get_arrayData(df)

                    if compare:
                        match1= re.search(f'{label}.*?EINHEIT_X\s"(.*?)"', dcm_data, re.S)
                        match2= re.search(f'{label}.*?EINHEIT_W\s"(.*?)"', dcm_data, re.S)

                        if match1:
                            if x_unit != match1.group(1): x_unit = match1.group(1)
                        if match2:
                            if l_unit != match2.group(1): l_unit = match2.group(1)

                    text += f"KENNLINIE {label} {x_num}\n"                + \
                            f"   FUNKTION {function}\n"                   + \
                            f'   EINHEIT_X "{x_unit}"\n'                  + \
                            f'   EINHEIT_W "{l_unit}"\n'                  + \
                            make_xText(list(df.iloc[0,1:]))+"\n"          + \
                            make_xText(list(df.iloc[1,1:]),"WERT")+"\n"   + \
                            "END\n\n"

                if is_listData(label,df, tab_pick):

                    label, val_unit, val_list = get_listData(label, df)

                    if compare:
                        match = re.search(f'{label}.*?EINHEIT_W\s"(.*?)"', dcm_data, re.S)[0]

                        if match:
                            if val_unit != match.group(1): val_unit = match.group(1)

                    text += f"FESTWERTEBLOCK {label} {len(val_list)}\n"   + \
                            f"   FUNKTION {function}\n"                   + \
                            f'   EINHEIT_W "{val_unit}"\n'                + \
                            make_xText(val_list, "WERT") + "\n"           + \
                            "END\n\n"

                if is_singleData(label, df, tab_pick, prepattern):

                    unit, value = get_singleData(label, df, prepattern)

                    if compare:
                        match = re.search(f'{label}.*?EINHEIT_W\s"(.*?)"', dcm_data, re.S)[0]

                        if match:
                            ref_unit = match.group(1)
                            if unit != ref_unit: unit = ref_unit

                    val_text = f"   WERT {value}\n"
                    if value.replace(" ", "").isalpha(): val_text = f'   TEXT "{value}"\n'

                    text += f"FESTWERT {label}\n"         + \
                            f"   FUNKTION {function}\n"   + \
                            f'   EINHEIT_W "{unit}"\n'    + \
                            val_text + \
                            "END\n\n"

    print("Step 5 / 6 has been completed, total runtime: " + \
    str(datetime.now()-startTime).split(".")[0]+" h/m/s\n")

#%%
    if os.path.exists("missing_dcm.json"): os.remove("missing_dcm.json")
    with open("missing_dcm.json","w") as f: json.dump(dcm_dict,f)

#%%
    if os.path.exists("missing.DCM"): os.remove("missing.DCM")
    with open('missing.DCM', 'x', encoding='utf-8') as f: f.write(text)

#%%
    print("Process has been completed. You may find output files in same directory.")
    print("Program runtime: " + str(datetime.now()-startTime).split(".")[0]+" h/m/s")

# ------------------------------------------------------------------------------------------ #

# Calling function:

import os, sys as sy

print("\nHello and welcome to the pdf analyzer mini tool.")
print("Please supply the path of CDR and PVER files when prompted.\n")

print("For very large pdf files (+30k pages) it is recommended to use at least 8 GB of RAM")
print("In default, program starts trying to read pdf file with 2 GB of RAM")
print("If the read was not succesfull, the tool increases RAM by 2GB and tries again.")
print("It is limited at 16 GB RAM usage. Hence the process is as 2-4-6-8...16 GB RAM")
v=input("If you want to start at a specific RAM amount please type. Otherwise press enter: ")

if v == "" : ram = 2

if v.isdigit():
    if int(v) < 18:
        ram = int(v)
    else:
        sy.exit("You have entered a unsupported value. Exiting the tool...")
else:
    sy.exit("Please only enter an integer. Exiting the tool...")

print("\nAfter the process, do you want to use a reference DCM file to compare units?")
print("If the units in output DCM file is different than reference, unit may be changed")
v=input("If you want to use reference file, enter DCM file path. Otherwise press enter: \n")

if v != "" :
    if os.path.isfile(v):
        dcm_path = v
    else:
        sy.exit("DCM file is not available at given path. Exiting the tool..")

cdr_path = input("\nCDR - Calibration Delivery Reporter file address: \n")
sw_path = input("\nPVER - Software Documentation file address: \n")

if cdr_path == "" or sw_path == "":
    sy.exit("One or more path has not been entered. Exiting the tool..")
if not os.path.isfile(cdr_path) or not os.path.isfile(sw_path):
    sy.exit("One or more file not available at given path. Exiting tool..")

print("\nProcess started, program can automatically close when finished. Please wait..")

if v == "" : dcm_missing(cdr_path, sw_path, ram)
if v != "" : dcm_missing(cdr_path, sw_path, ram, dcm_path)

# ------------------------------------------------------------------------------------------ #
