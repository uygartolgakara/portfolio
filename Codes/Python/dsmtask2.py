# -*- coding: utf-8 -*-
"""
Produced on Sat Sep  9 16:17:47 2023
@author: KUY3IB
"""

# %% Import Modules

import xml.etree.ElementTree as ET
from collections import defaultdict
from os import listdir
from os.path import join
from openpyxl import load_workbook
import win32com.client as win32


# %% Hardcoded Variables

# no lamp activation
NLA = [
    0,
    1,
    2,
    8,
    11,
    15,
    16,
    19
]

# continuous lamp support
CLS = [
    3,
    4,
    5,
    6,
    7,
    9,
    10,
    13,
    14,
    17,
    18,
    20,
    21
]

# not used
NU = [
    12,
    22,
    23,
    25,
    26,
    27,
    28,
    29
]

# blinking lamp support
BLS = [
    24
]

# %%


class CDFX_Tool:
    """Provides 2 functions to check FID or DFC in CDFX file content."""

    def __init__(self, cdfx_path: str):
        """
        Initialize parameters and prepare functions.

        Args:
            cdfx_path (str): Path of cdfx file in raw string format.

        Returns:
            None.

        """
        # Define global variables for debugging
        global cdfx_dict
        # global sw_instance_list
        # global dmask2_list
        # global dclass_list
        # global dfid_list
        # global dmask2_vals
        # global dclass_vals
        # global dfid_vals
        global dfc_dict
        global fid_dict
        global xlsm_files
        global xlsm_paths
        global mail_text

        # Read CDFX file and make nested dictionary
        cdfx_dict = self.xml_file_to_dict(cdfx_path)

        # Gather sw instance list from cdfx dictionary
        sw_instance_list = self.find_sub_list(cdfx_dict, "SW-INSTANCE")

        # Gather dfes class related lists
        dclass_list = self.find_starting(
            sw_instance_list, "SHORT-NAME", "DFES_Cls.")

        # Gather disable mask 2 related lists
        dmask2_list = self.find_starting(
            sw_instance_list, "SHORT-NAME", "DFC_DisblMsk2.")

        # Gather finh fid related lists
        dfid_list = self.find_starting(
            sw_instance_list, "SHORT-NAME", "DINH_FId.")

        # Find dfes class values of DFCs
        dclass_vals = self.disect_list(dclass_list)

        # Find disable mask 2 values of DFCs
        dmask2_vals = self.disect_list(dmask2_list)

        # Find contained fid values of DFCs
        dfid_vals = self.disect_listoflists(dfid_list)

        # Make main dictionary with DFC keys
        self.dfc_dict = self.make_dfc_dict(dmask2_vals, dclass_vals, dfid_vals)

        # Remove item if unused dfc is present
        if "DFC_Unused" in self.dfc_dict:
            del self.dfc_dict["DFC_Unused"]

        # Remove item if unused dfc is present
        if "DSQ_Unused" in self.dfc_dict:
            del self.dfc_dict["DSQ_Unused"]

        # Make main dictionary with FId keys
        self.fid_dict = self.make_fid_dict()

        # Gather path of Excel file directory
        dir_path = input("Enter path for directory: ")

        # Gather list of xlsm files
        xlsm_files = listdir(dir_path)
        xlsm_files = [file for file in xlsm_files if file.endswith(".xlsm")]
        xlsm_paths = [join(dir_path, file) for file in xlsm_files]

        # Construct mail text string
        mail_text = []

        # Start the loop for iteration
        for xlsm_path in xlsm_paths:

            wb = load_workbook(xlsm_path, read_only=True)
            ws = wb["FID_Sicht"]

            for i in range(10, ws.max_row + 1):

                print(f"{i}/{ws.max_row+1}")

                a_val = ws[f"A{i}"].value  # fid value
                c_val = ws[f"C{i}"].value
                d_val = ws[f"D{i}"].value
                e_val = ws[f"E{i}"].value
                f_val = ws[f"F{i}"].value

                if (
                        (a_val == "" or a_val is None or a_val == "None")
                        and (c_val == "" or c_val is None or c_val == "None")
                        and (d_val == "" or d_val is None or d_val == "None")
                        and (e_val == "" or e_val is None or e_val == "None")
                        and (f_val == "" or f_val is None or f_val == "None")
                    ):
                    break

                if (
                        (c_val == "" or c_val is None or c_val == "None")
                        and (d_val == "" or d_val is None or d_val == "None")
                    ):

                    if a_val in list(self.fid_dict.keys()):
                        if e_val not in self.fid_dict[a_val]:
                            mail_text.append(f"{a_val} needs to be added to {e_val}")
                    else:
                        pass
                        # mail_text.append(f"{a_val} needs to be added to {e_val}")

                if (
                        (e_val == "" or e_val is None or e_val == "None")
                        and (f_val == "" or f_val is None or f_val == "None")
                    ):

                    if c_val in list(self.dfc_dict.keys()):
                        if a_val in self.dfc_dict[c_val]:
                            mail_text.append(f"{a_val} needs to be removed from {c_val}")

        # Convert list of lines to a single HTML-compatible string
        html_body = '<br>'.join(mail_text)

        # Make mail
        self.send_outlook_mail(
            recipients="recipient1@example.com",
            subject="Test Subject",
            body=html_body,
            # attachment_path="path_to_file.txt"  # Replace with your file's path or remove if not needed
        )














    def xml_file_to_dict(self, file_path: str) -> dict:
        """
        Prepare a nested dictionary from cdfx file.

        Args:
            file_path (str): Path of a cdfx file.

        Returns:
            dict: Nested dictionary from cdfx content.

        """
        tree = ET.parse(file_path)
        return {tree.getroot().tag: self.etree_to_dict(tree.getroot())}

    def etree_to_dict(self, t) -> dict:
        """
        Support nested dictionary preparation.

        Args:
            t (TYPE): Tag of root level.

        Returns:
            d (dict): Nested dictionary.

        """
        d = {t.tag: {} if t.attrib else None}
        children = list(t)
        if children:
            dd = defaultdict(list)
            for dc in map(self.etree_to_dict, children):
                for k, v in dc.items():
                    dd[k].append(v)
            d = {t.tag: {k: v[0] if len(v) == 1 else v for k, v in dd.items()}}
        if t.attrib:
            d[t.tag].update(('@' + k, v) for k, v in t.attrib.items())
        if t.text:
            text = t.text.strip()
            if children or t.attrib:
                if text:
                    d[t.tag]['#text'] = text
            else:
                d[t.tag] = text
        return d

    def find_sub_list(self, main_dict: dict, target: str) -> dict:
        """
        Find sub dictionary which has lists as values in nested dictionary.

        Args:
            main_dict (dict): Main nested dictionary to be iterated.
            target (str): Key of the sub dictionary for searching.

        Returns:
            dict: Sub dictionary fround inside nested dictionary.

        """
        if isinstance(main_dict, dict):
            for key, value in main_dict.items():
                if key == target:
                    return value
                else:
                    result = self.find_sub_list(value, target)
                    if result:
                        return result

    def find_starting(self, main_list: list, key: str, title: str) -> list:
        """
        Find list items starting with a title string.

        Args:
            main_list (list): List with same format items to be iterated.
            key (str): Key of the dictionaries of the searched value.
            title (TYPE): Serached substring for starting of list items.

        Returns:
            list: Conditionally satisfied item list as sub-list.

        """
        sub_list = []
        for item in main_list:
            if item[key].startswith(title):
                sub_list.append(item)
        return sub_list

    def disect_list(self, item_list: list) -> dict:
        """
        Use data in list of dictionary to make a new filtered dictionary.

        Args:
            item_list (list): List of one format dictionaries.

        Returns:
            dict: Dictionary of DFC names with physical values.

        """
        output_dict = {}

        for item in item_list:
            dfc_name = item["SHORT-NAME"].split(".")[1]
            value = item["SW-VALUE-CONT"]["SW-VALUES-PHYS"]["V"].split(".")[0]
            output_dict[dfc_name] = value

        output_dict = {key.rsplit(
            '_C', 1)[0]: value for key, value in output_dict.items()}

        return output_dict

    def disect_listoflists(self, item_list: list):
        """
        Similar to dist_list function, but used for list of lists.

        Args:
            item_list (list): List of list of lists of dictionaries.

        Returns:
            output_dict (TYPE): Dictionary of FId names with physical values.

        """
        output_dict = {}

        for item in item_list:
            dfc_name = item["SHORT-NAME"].split(".")[1]
            output_dict[dfc_name] = []

            sub_list = item["SW-VALUE-CONT"]["SW-VALUES-PHYS"]["VT"]
            for sub in sub_list:
                sub_name = sub["#text"]
                if sub_name == "FId_Unused":
                    continue
                output_dict[dfc_name].append(sub_name)

        output_dict = {key.rsplit(
            '_CA', 1)[0]: value for key, value in output_dict.items()}

        return output_dict

    def make_dfc_dict(self, dict1: dict, dict2: dict, dict3: dict) -> dict:
        """
        Join disable mask, disable class and fid dictionaries.

        Args:
            dict1 (dict): Disable mask 2 dictionary.
            dict2 (dict): Disable class dictionary.
            dict3 (dict): DFC-FID names dictionary.

        Returns:
            dict: Combined data from input datas.

        """
        dfc_dict = {}  # Make an empty dictionary

        # Add all key values as DFC and DSQ values
        all_dfcs = set(
            list(dict1.keys())
            + list(dict2.keys())
            + list(dict3.keys())
        )

        # Using DFC names as keys gather disable mask and disable class
        for dfc in all_dfcs:
            nested_dict = {
                'dict1': dict1.get(dfc, None),
                'dict2': dict2.get(dfc, None),
                'dict3': dict3.get(dfc, None)
            }
            dfc_dict[dfc] = nested_dict

        return dfc_dict

    def make_fid_dict(self) -> dict:
        """
        Make a dictionary with FId keys and DFC list as values.

        Returns:
            dict: Dictionary with FId keys.

        """
        transformed_dict = {}

        for key, value in self.dfc_dict.items():
            fid_list = value["dict3"]
            if fid_list is None:
                continue
            if len(fid_list) == 0:
                continue
            for item in fid_list:
                if item not in list(transformed_dict.keys()):
                    transformed_dict[item] = [key]
                else:
                    transformed_dict[item].append(key)
        return transformed_dict

    def check_dfc(self, dfc_name: str) -> str:
        """
        Check a situation of a given DFC name.

        Args:
            dfc_name (str): Name of a DFC or DSQ to be checked.

        Returns:
            str: State of DFC

        """
        # Gather all available dfc names
        dfc_names = list(self.dfc_dict.keys())

        # If DFC name is given out ot CDFX
        if dfc_name not in dfc_names:
            return ["Missing", "Missing", "DFC not available"]

        # If dfc name is DSQ, assume it is okay
        if dfc_name.startswith("DSQ_"):
            return ["Missing", "Missing", "Okay"]

        # Gather related dfc data
        dfc_data = self.dfc_dict.get(dfc_name)

        # Disect and check dfes class and disable mask
        dcls = dfc_data["dict2"]
        dmsk = dfc_data["dict1"]

        # Conditionally return if one value is missing
        if dcls is None and dmsk is not None:
            return ["Missing", dmsk, "DFES Class not available"]
        if dcls is not None and dmsk is None:
            return [dcls, "Missing", "Disable Mask not available"]
        if dcls is None and dmsk is None:
            return ["Missing", "Missing",
                    "DFES Class and Disable Mask not available"]

        # Check if both values are scalar
        try:
            dmsk = int(dmsk)
            dcls = int(dcls)
        except Exception:
            return ["Scalar Error", "Scalar Error",
                    "DFES Class or Disable Mask is not scalar"]

        # Continuous Lamp Support case
        if dcls in CLS:
            return [dcls, dmsk, "Okay"]

        # No Lamp Activation case
        if dcls in NLA:
            # gather 16 bit translation of dmask
            dmsk_16bit = bin(dmsk)[2:].zfill(16)
            # check if 2nd bit to right is 1
            condition1 = dmsk_16bit[-2] == "1"
            # check if all bits are 1
            condition2 = all(bit == '1' for bit in dmsk_16bit)
            # check if one condition is satisfied
            if condition1 or condition2:
                return [dcls, dmsk, "Okay"]
            else:
                return [dcls, dmsk, "Not Okay"]

        # Not Used case
        if dcls in NU:
            return [dcls, dmsk, "Suspected"]

        # Unknown case
        return [dcls, dmsk, "Error"]

    def check_fid(self, fid_name: str) -> str:
        """
        Check a situation of a given FId name.

        Args:
            fid_name (str): Name of FId to be checked.

        Returns:
            str: State of FID.

        """
        # Gather all available fid names
        fid_names = list(self.fid_dict.keys())

        # If FId name is given out ot CDFX
        if fid_name not in fid_names:
            return ["Missing", "Missing", "FID not available"]

        # Gather related fid data
        dfc_list = self.fid_dict.get(fid_name)

        # DFC results
        result_dict = {}

        # Check all related DFC's to FID
        for dfc in dfc_list:
            result_list = self.check_dfc(dfc)
            result_dict[dfc] = result_list

        # Start of FID related DFC data printing
        print("\n----- FID connections -----\n")

        # Iterate fid related dfc data and print
        for index, dfc in enumerate(dfc_list):
            print(
                f"{dfc} : {result_dict[dfc][2]}\n"
                f"DFES Class: {result_dict[dfc][0]}\n"
                f"Disable Mask: {result_dict[dfc][1]}\n"
            )

        # Start of FID data printing
        print("\n----- FID decision -----\n")

        # If all results are okay, fid is okay
        if all(value[2] == 'Okay' for value in result_dict.values()):
            print(f"{fid_name} : Okay")
            return result_dict

        # If any results is not okay, fid is not okay
        if any(value[2] == 'Suspected' for value in result_dict.values()):
            print(f"{fid_name} : Suspected")
            return result_dict

        # If any results is not okay, fid is not okay
        if any(value[2] == 'Not Okay' for value in result_dict.values()):
            print(f"{fid_name} : Not Okay")
            return result_dict

        # In other cases
        print(f"{fid_name} : Not Okay")
        return result_dict

    def send_outlook_mail(self, recipients, subject, body, attachment_path=None):
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)

        mail.To = recipients
        # mail.CC = "cc_email@example.com"  # Uncomment and set if needed
        mail.Subject = subject
        mail.HTMLBody = body

        # Add an attachment if provided
        if attachment_path:
            mail.Attachments.Add(attachment_path)

        mail.Display()  # .Send() to send or .Save() to save to drafts

    # # Example usage
    # send_outlook_mail(
    #     recipients="recipient1@example.com",
    #     subject="Test Subject",
    #     body="<h1>This is a test email sent from Python.</h1>",
    #     attachment_path="path_to_file.txt"  # Replace with your file's path or remove if not needed
    # )

if __name__ == "__main__":

    # Ask user for path of cdfx file
    cdfx_path = input("\nEnter path of CDFX file: ")

    # Make class object
    cdfx_tool = CDFX_Tool(cdfx_path)

    # # Test case for checking DFC
    # dfc_name = "DFC_AfterrunBegin"
    # result = cdfx_tool.check_dfc(dfc_name)
    # print(dfc_name + ": " + result)

    # # Test case for checking DFC
    # dfc_name = "DFC_AirbgCrCtlDisbl"
    # result = cdfx_tool.check_dfc(dfc_name)
    # print(dfc_name + ": " + result)

    # # Test case for checking FId
    # fid_name = "FId_ACCompAirT"
    # result = cdfx_tool.check_fid(fid_name)
    # print(fid_name + ": " + result)

    # # Test case for checking FId
    # fid_name = "FId_ACCtl_tEnv"
    # result = cdfx_tool.check_fid(fid_name)
    # print(fid_name + ": " + result)

    # # Test case for checking FId
    # fid_name = "FId_ClthDiagErr"
    # result = cdfx_tool.check_fid(fid_name)
    # print(fid_name + ": " + result)
