# -*- coding: utf-8 -*-

"""
Produced on Fri Aug 25 17:06:11 2023.

@author: Uygar Tolga Kara
"""

# %% Import Modules

from time import time as tm
from os.path import basename
from openpyxl.styles import Alignment
from pandas import read_csv, DataFrame
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, worksheet
from tkinter import filedialog, messagebox, Tk
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.formatting.formatting import ConditionalFormattingList

# %% Hardcoded Variables

HEADER = [
    "Action:",
    "DFC:",
    "Inhibit:",
    "Limit:",
    "Problems:"
]

PROBLEM = {
    "inhibit already present": "Included in SW Config",
    "inhibit not found": "Requirement Invalid",
    "dfc/fid not found": "DFC/DSQ/FId Invalid",
    "dfc/fid invalid": "DFC/DSQ/FId Invalid",
    "dfc invalid": "DFC/DSQ/FId Invalid",
    "fid invalid": "DFC/DSQ/FId Invalid",
    "dsq invalid": "DFC/DSQ/FId Invalid",
    "dfc not found": "DFC/DSQ/FId Invalid",
    "fid not found": "DFC/DSQ/FId Invalid",
    "dsq not found": "DFC/DSQ/FId Invalid",
    "no free sst available": "No free SST available"
}

VALIDATION_A = {
    "Included in SW Config": ["included in sw config", "#FF99FF99"],
    "Requirement Invalid": ["requirement invalid", "#FFEAEAEA"],
    "DFC/DSQ/FId Invalid": ["dfc/dsq/fid invalid", "#FFEAEAEA"],
    "No free SST available": ["no free sst available", "#FFFF6699"]
}

VALIDATION_C = {
    "Add Inhibit": ["add inhibit", "#FF00B050"],  # red
    "Remove Inhibit": ["remove inhibit", "#FFFF0000"]  # green
}

# %% Class Area


class DsmTool:
    """Uses a logging file and a function file to make an edited file."""

    def __init__(self, seperator=";"):
        """
        Prepare input files and complete operations.

        Args:
            seperator (TYPE, optional): Log file seperator. Defaults to ";".

        Returns:
            None.

        """
        # Make global variables
        global logging_path
        global function_path
        global df, wb, ws

        # start timer
        start_time = tm()

        # inform user
        print("\nHello, please select the files when prompted\n")

        # Gather path of log file
        print("Please browse and select the logging file")
        logging_path = self.browse_file("csv")

        # Gather path of function file
        print("Please browse and select the function file")
        function_path = self.browse_file("xlsx")

        # Print file names of user selected log and function files
        print(f"\nSelected logging file: {basename(logging_path)}")
        print(f"Selected function file: {basename(function_path)}\n")

        # Read csv formatted log file as dataframe
        df = read_csv(logging_path, sep=seperator)

        # Modify dataframe with pandas operations
        df = self.prepare_data(df)

        # Read xlsx formatted function file as workbook
        wb = load_workbook(function_path)
        ws = wb.worksheets[0]

        # Reset worksheet file's data validation
        ws = self.prepare_worksheet(ws)

        # Make data validation objects
        self.dv_A = self.construct_validation(VALIDATION_A)
        self.dv_C = self.construct_validation(VALIDATION_C)

        # Add data validation to worksheet
        ws.add_data_validation(self.dv_A)
        ws.add_data_validation(self.dv_C)

        # Reset values, validation, coloring on worksheet
        ws = self.update_worksheet(ws, df)

        # Start loading animation
        print("Excel file is being saved. Please wait...")

        # Save the workbook
        wb.save("edited.xlsx")

        # Calculate runtime
        runtime = tm() - start_time

        # Disect and print runtime
        self.calculate_time(runtime)

    def browse_file(self, filetype: str) -> str:
        """
        Browses file path based on related with file type.

        Args:
            filetype (str): Type of the file to be searched. (csv or xlsx)

        Raises:
            RuntimeError: Exit error to be shown when file is not submitted.

        Returns:
            str: Path of the selected file on browse window.

        """
        # Make and hide the browse window
        window = Tk()
        window.withdraw()

        # Set filetype variable
        match filetype:
            case "csv":
                filetypelist = [("CSV files", "*.csv")]
            case "xlsx":
                filetypelist = [("Excel files", "*.xlsx")]
            case _:
                raise RuntimeError("Wrong file type")

        # Request user to browse a file
        filepath = filedialog.askopenfilename(filetypes=filetypelist)

        # If file is not given, ask for decision to continue or cancel
        while not filepath:

            decision = messagebox.askquestion(
                "No File Selected",
                "You did not select required file.\n"
                "Would you like to browse again?",
                icon='warning', default='yes', type='yesno')

            # Do related task based on decision
            match decision:
                case "yes":
                    filepath = self.browse_file(filetype)
                case "no":
                    raise RuntimeError("Application terminated")

        return filepath

    def prepare_data(self, df: DataFrame) -> DataFrame:
        """
        Preprocess data inside dataframe.

        Args:
            df (DataFrame): Dataframe read from csv file

        Returns:
            DataFrame: Dataframe with preprocessed data

        """
        # Gather column list of dataframe
        df.columns = list(range(df.shape[1]))

        # Find row index of hardcoded list
        for i, row in df.iloc[::-1].iterrows():
            if all(header in row.values for header in HEADER):
                row_index = i
                break

        # Gather sub dataframe
        df = df[row_index:]

        # Reset row indices
        df = df.reset_index(drop=True)

        # Use first row as header
        df.columns = df.iloc[0]

        # Drop 2 rows of dataframe
        df = df.drop(index=[0, 1])

        # Reset row indices
        df = df.reset_index(drop=True)

        # Drop columns other than header list
        df = df[HEADER]

        # Drop nan columns
        df = df.dropna(axis=1, how='all')

        return df

    def prepare_worksheet(self, ws: worksheet):
        """
        Perform types of formatting on worksheet.

        Args:
            ws (worksheet): Function file 1st worksheet.

        Returns:
            ws (TYPE): New format applied worksheet.

        """
        # Remove all conditional formatting
        ws.conditional_formatting = ConditionalFormattingList()

        # Remove all data validation rules
        ws.data_validations = DataValidationList()

        # Set zoom level to 90%
        ws.sheet_view.zoomScale = 90

        # Set the width of column 'A' to 25 units
        ws.column_dimensions['A'].width = 25

        # Conditional formatting of a cell value
        if str(ws["A4"].value) == "DINH":
            ws["A4"].value = ""
            ws["A3"].value = "DINH"
            ws["A3"].font = Font(bold=True)

        return ws

    def construct_validation(self, options: list) -> DataValidation:
        """
        Make data validation object from list of options.

        Args:
            options (list): Potential values an excel cell may have.

        Returns:
            DataValidationList: Data validation object for excel file.

        """
        # Join options in a string
        options_str = ",".join(options)

        # Make data validation list object
        dv = DataValidation("list", f'"{options_str}"', allow_blank=True)

        return dv

    def update_worksheet(self, ws: worksheet, df: DataFrame) -> worksheet:
        """
        Iterate and modify rows in worksheet.

        Args:
            ws (worksheet): Function workbook's 1st worksheet.
            df (DataFrame): Dataframe from logging file content.

        Returns:
            worksheet: Modified function workbook's 1st worksheet.

        """
        # Enter a loop of rows in excel
        for index in range(5, ws.max_row + 1):

            # Gather values in columns
            a_val = ws[f'A{index}'].value
            b_val = ws[f'B{index}'].value
            c_val = ws[f'C{index}'].value
            d_val = ws[f'D{index}'].value
            e_val = ws[f'E{index}'].value

            # Skip row if only one column is filled
            if None in [b_val, c_val, d_val, e_val]:
                continue

            # Check A column value as case
            for key, value in VALIDATION_A.items():
                if str(a_val).strip().lower() == value[0]:
                    ws[f'A{index}'].value = a_val = key
                    break
            else:
                ws[f'A{index}'].value = a_val = ""

            # Check C column value as case
            for key, value in VALIDATION_C.items():
                if str(c_val).strip().lower() == value[0]:
                    ws[f'C{index}'].value = c_val = key
                    break
            else:
                ws[f'C{index}'].value = c_val = ""

            # Apply text alignment on A column
            ws[f'A{index}'].alignment = Alignment(horizontal='left')

            # Produce string values of values
            b_str = str(b_val).strip().lower()
            c_str = str(c_val).strip().lower()
            d_str = str(d_val).strip().lower()
            e_str = str(e_val).strip().lower()

            # Compare dataframe and worksheet
            for idx in range(df.shape[0]):
                check1 = b_str == str(df.iloc[idx, 1]).strip().lower()
                check2 = c_str == str(df.iloc[idx, 0]).strip().lower()
                check3 = d_str == str(df.iloc[idx, 2]).strip().lower()
                check4 = e_str == str(df.iloc[idx, 3]).strip().lower()

                # If all conditions are satisfied, update value
                if all([check1, check2, check3, check4]):
                    problem = str(df.iloc[idx, 4]).lower()
                    if problem in list(PROBLEM.keys()):
                        translation = PROBLEM[problem]
                        ws[f'A{index}'].value = translation
                        break

            # Apply data validation for A and C
            self.dv_A.add(ws[f'A{index}'])
            self.dv_C.add(ws[f'C{index}'])

            # Add conditional formatting on A column
            for key, value in VALIDATION_A.items():
                color_hex = value[1][1:]
                fill = PatternFill(
                    start_color=color_hex,
                    end_color=color_hex,
                    fill_type="solid")
                formula = [f'$A{index}="{key}"']
                rule = FormulaRule(formula=formula, fill=fill)
                ws.conditional_formatting.add(f'A{index}', rule)

            # Add conditional formatting on C column
            for key, value in VALIDATION_C.items():
                color_hex = value[1][1:]
                font = Font(color=color_hex)
                formula = [f'$C{index}="{key}"']
                rule = FormulaRule(formula=formula, font=font)
                ws.conditional_formatting.add(f'C{index}', rule)

        return ws

    def calculate_time(self, runtime: float):
        """
        Calculate hours, minutes, seconds from runtime and print.

        Args:
            runtime (float): Seconds value gathered from time module.

        Returns:
            None.

        """
        # Calculate hours, minutes and seconds
        hours, remainder = divmod(runtime, 3600)
        minutes, seconds = divmod(remainder, 60)

        # Produce time texts
        h_text = f"{int(hours)} hours"
        m_text = f"{int(minutes)} minutes"
        s_text = f"{int(seconds)} seconds"

        # Print conditional time
        if hours > 0:
            if minutes > 0:
                print(f"Runtime: {h_text}, {m_text} and {s_text}")
            else:
                print(f"Runtime: {h_text} and {s_text}")
        else:
            if minutes > 0:
                print(f"Runtime: {m_text} and {s_text}")
            else:
                print(f"Runtime: {s_text}")


if __name__ == "__main__":

    # Make class object
    dsm_tool = DsmTool()
