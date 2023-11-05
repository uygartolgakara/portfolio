# Author : KUY3IB - Developed on Tue Aug  9 10:44:05 2022

#%% Import Libraries

from packaging import version
#import packaging as pg
import openpyxl as op
import PyPDF2 as p2
import shutil as sh
import os as os
import re as re

#%% Get User Inputs

nr_dir = input("Neutral Directory: ")
sw_dir = input("SW Directory: ")
fo_num = input("Folder Numbers: ")
ou_dir = input("Output Directory: ")

#%% Reset and Make Output Folder

ou_dir = os.path.join(ou_dir,"Output")
if os.path.exists(ou_dir): sh.rmtree(ou_dir)
if not os.path.exists(ou_dir): os.makedirs(ou_dir)

#%% Find Target File Paths

nr_list = os.listdir(nr_dir)
sw_list = os.listdir(sw_dir)

fo_list =  fo_num.strip().split()

nr_name = [name for name in nr_list if name.startswith(tuple(fo_list))]
nr_path = [os.path.join(nr_dir, name, "10-CDR") for name in nr_name]

#%% Main Loop Per Neutral Folder Path

sw_dict = {}

for base_index, base_path in enumerate(nr_path):

    fol_name = os.listdir(base_path)
    rep_name = [name for name in fol_name if "CalibrationDeliveryReporter" in name]

    if len(rep_name) == 0: continue

    rep_path = os.path.join(base_path, rep_name[0])

    wb = op.load_workbook(rep_path)
    ws = wb['Calibration details']

    #%% Secondary Loop per CDR Excel File

    for i in range(ws.min_row+2, ws.max_row+1):
        if ws.row_dimensions[i].hidden == False and\
            ws[f"F{i}"].value == "missing" or ws[f"H{i}"].value == "initial calibration":

            funct = ws[f"A{i}"].value
            label = ws[f"D{i}"].value
            newvs = ws[f"C{i}"].value

            if funct == ""   or label == ""  : continue
            if funct is None or label is None: continue

            sw_name = [name for name in sw_list if name.startswith(funct)]
            sw_name = [name for name in sw_name if newvs in name]
            if len(sw_name) == 0: continue

            sw_ver = [name.split(funct + "_")[1] for name in sw_name]
            sw_ver = [nm.split("_")[0] + "." + nm.split("_")[1] for nm in sw_ver]

            latest_version = max(sw_ver, key=version.parse)
            latest_index = sw_ver.index(latest_version)

            sw_name = sw_name[latest_index]
            sw_path = os.path.join(sw_dir, sw_name)

            if sw_name in list(sw_dict.keys()): text = sw_dict[sw_name]

            # ---------------------------------------------------------------------

            if sw_name not in list(sw_dict.keys()):

                pdfFileObj = open(sw_path, 'rb')
                pdfReader = p2.PdfFileReader(pdfFileObj)
                pageNum = pdfReader.numPages

                text = ""
                for i in range(pageNum): text += pdfReader.getPage(i).extractText()

            sw_dict[sw_name] = text

            # ---------------------------------------------------------------------

            regex_one = ".{1,750}?Start Value : .*?\]|Standart Value : .*?\]"
            match = re.search(label + regex_one, text, re.S)
            if match is None:
                ws[f"J{i}"].value = "No Value is Available in SW Document"
                continue

            label_text  = text[match.start():match.end()]
            # print(text[match.start(): match.end()])

            regex_two = "Start Value : .*?\]|Standart Value : .*?\]"
            label_value = re.findall(regex_two, label_text, flags=re.S)[0]
            # print(label_value)

            ws[f"J{i}"].value = label_value

    # -----------------------------------------------------------------------------
    save_path = os.path.join(ou_dir,nr_name[base_index])
    os.makedirs(save_path)
    wb.save(os.path.join(save_path,rep_name[0])); wb.close()
