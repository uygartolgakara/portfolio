# -*- coding: utf-8 -*-
"""
Produced on Wed Sep  6 15:04:58 2023
@author: KUY3IB
"""
import openpyxl as op

wb1 = op.load_workbook(
    r"C:\Users\KUY3IB\Desktop\Test\Copy of Bosch eskalasyo.xlsx")
ws1 = wb1["eskaslasyon uygulama"]

wb2 = op.load_workbook(
    r"C:\Users\KUY3IB\Desktop\Test\Copy of 1. 2023 Ocak-Temmuz_Bosch STS Özet (002).xlsx")
ws2 = wb2["Ford_PS"]

wb3 = op.load_workbook(
    r"C:\Users\KUY3IB\Desktop\Test\2023_RBTR_SO_Turnover .xlsx")
ws3 = wb3["Unit "]

for i in range(5, ws1.max_row + 1):
    for ii in range(6, ws3.max_row + 1):
        if ws1[f"A{i}"].value == ws2[f"C{ii}"].value:
            ws1[f"B{i}"].value = ws2[f"B{ii}"].value
            ws1[f"E{i}"].value = ws2[f"J{ii}"].value
            continue

for i in range(5, ws1.max_row + 1):
    for ii in range(5, ws3.max_row + 1):
        if ws1[f"B{i}"].value == ws3[f"D{ii}"].value:
            ws1[f"D{i}"].value = ws3[f"W{ii}"].value
            continue


wb1.save("modified.xlsx")

# wb1 = op.load_workbook("modified.xlsx")
# ws1 = wb1["eskaslasyon uygulama"]


# ---------------------------------------------------------------------------
