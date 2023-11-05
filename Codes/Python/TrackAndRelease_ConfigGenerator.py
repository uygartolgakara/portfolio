#%% Import libraries
import openpyxl as op  # To read excel,outline level
import shutil  as sh # To delete files if required
import pandas as pd # To read sheet as dataframe
import numpy as np # To search string in array
import json as js # To make json file at end
import sys as sy # To message and exit app
import os as os # To access system files
import re

#%% Import modules for Logging
import datetime as dt # To save time marks
from openpyxl.styles.borders import Border, Side # To format log entries
from openpyxl.styles import Font, Color, Fill, PatternFill # To format log
pd.options.mode.chained_assignment = None # To avoid getting copy warnings

#%% All functions in first usage order in script- Scroll past to code
def start_logger(log_path:str):
    """ Makes blank log Excel file with headers and saves """
    log_wb = op.Workbook()
    log_ws = log_wb.active
    label1, label2, label3 = "Timestamp" , "Message" , "Type"
    label4, label5 = "Consequence/Result", "Suggested Action"
    log_ws.append([label1, label2, label3, label4, label5])
    column_widths = {"A" : 20, "B": 60, "C": 10, "D" : 60, "E": 60}
    for c,w in column_widths.items(): log_ws.column_dimensions[c].width = w
    headers = ["A1", "B1", "C1", "D1", "E1"]
    for tab in headers:
        log_ws[tab].font = Font(bold = True)
        log_ws[tab].fill = PatternFill(start_color = "BCBCBC", fill_type = "solid")
        log_ws[tab].border = Border(bottom = Side(style = 'thick'))
    log_wb.save(log_path)
def record_log(log_path:str, label2:str, label3:str, label4:str, label5:str):
    """ Adds a line record_log on the log file """
    wb = op.load_workbook(log_path)
    ws = wb.active
    max_row = ws.max_row + 1
    timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([timestamp, label2, label3, label4, label5])
    tab = ["A","B","C","D","E"]
    row_area = [col+str(max_row) for col in tab]
    for loc in row_area:
        if label3 == "INFO":
            ws[loc].fill = PatternFill(start_color = "93c47d", fill_type = "solid")
        elif label3 == "WARNING":
            ws[loc].fill = PatternFill(start_color = "ffd966", fill_type = "solid")
        elif label3 == "ERROR":
            ws[loc].fill = PatternFill(start_color = "e06666", fill_type = "solid")
    wb.save(log_path)
def filter_worksheets(worksheets:list, keywords:list, exceptions:list) -> list:
    """ Filter worksheets containing keywords if not an exception """
    for keyword in keywords:
        # Make variations to keyword to escape case sensitivity
        keys = [keyword, keyword.lower(), keyword.title(), keyword.upper()]
        for key in keys:
            for ws in worksheets:
                if key in ws and ws not in exceptions:
                    worksheets.remove(ws); break
def make_merged_cell_reference(sheet:object) -> dict:
    """ Support function for unmerging cells, save values and positions"""
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = op.utils.range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        merged_lookup[str(cell_group)] = top_left_cell_value
    return merged_lookup
def unmerge_worksheet_copy_top_left_value(wb:object, worksheet_name:str):
    """ Main function for unmerging operation, unmerge and change values """
    sheet = wb[worksheet_name]
    lookup = make_merged_cell_reference(sheet)
    cell_group_list = lookup.keys()
    for cell_group in cell_group_list:
        min_col, min_row, max_col, max_row = op.utils.range_boundaries(str(cell_group))
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, \
                                   max_col=max_col, max_row=max_row):
            for cell in row: cell.value = lookup[cell_group]
def drop_comment_column(ws_dict:dict, c_label:list):
    """ If a dataframe has a comment column, drop comment column """
    for df in ws_dict.values():
        da = df.values # Make an array for more easier search
        for comment in c_label:
            if comment in da:
                c_index = list(np.where(da == comment)[1])
                df.drop(c_index, axis=1, inplace=True); break
def gather_project_names(ws_dict:dict, p_label:list) -> list:
    """ If a dataframe has project labels, gather project names"""
    for ws, df in ws_dict.items():
        da = df.values # Make an array for more easier search
        for project in p_label:
            if project in da:
                row_index = np.where(da == project)[0][0]
                col_index = list(np.where(da == project)[1])
                return list(da[row_index + 1, col_index]), project
def gather_project_choice_sheets(ws_dict:dict, project_name:list) -> list:
    """ Gathers worksheet names which contain project columns """
    return [ws for ws,df in ws_dict.items() if project_name[0] in df.values]
def drop_empty_rows_and_columns(d_frame:object):
    """ In dataframe drops row and column if empty, and reset index """
    # Drop fully empty rows and reset index
    d_frame.dropna(axis=0, how='all', inplace=True)
    d_frame.reset_index(drop=True, inplace=True)
    # Drop fully empty columns and reset index
    d_frame.dropna(axis=1, how='all', inplace=True)
    d_frame.columns = range(d_frame.shape[1])
def find_outline_levels(ws:object) -> int:
    """ Find outline levels of a worksheet """
    row_level = col_level = 0 # reset
    # Find maximum outline level for the rows
    for index in range(ws.min_row, ws.max_row):
        new_level = ws.row_dimensions[index].outline_level
        if new_level > col_level: col_level = new_level
    col_level = col_level + 1 # Names as col_level because vertical
    # Find maximum outline level for the columns
    for index in range(ws.min_column, ws.max_column):
        tab = ws.cell(ws.min_row, index).column_letter
        new_level = ws.column_dimensions[tab].outline_level
        if new_level > row_level: row_level = new_level
    row_level = row_level + 1 # Named as row_level because horizontal
    return row_level, col_level
def drop_duplicate_rows(d_frame:object, row_level:int):
    """ If you are not sure, do not use df.drop_duplicates(); use this """
    if row_level > 0:
        for index in range(row_level):
            if d_frame.iloc[index].tolist() == d_frame.iloc[index+1].tolist():
                d_frame.drop(index, inplace=True)
                d_frame.reset_index(drop=True, inplace=True); break
            if index+1 == row_level: break
def identify_duplicate(df:object, row_level:int, col_level:int, w_sheet:str):
    """ If column in last index has duplicate values, add information to identify """
    ref_list = df.copy().iloc[row_level:, col_level-1].tolist()
    for idx, var in enumerate(ref_list):
        if var is not None and ref_list.count(var) > 1:
            df.iloc[idx+row_level,col_level-1] = str(df.iloc[idx+row_level,col_level-1])+"$"+str(idx)
        if var is None:
            df.iloc[idx+row_level,col_level-1] = str(df.iloc[idx+row_level,col_level-1])+"$"+str(idx)
def disect_dataframe(d_frame:object, project_name:list, row_level:int) -> list:
    """ Based on choices, disect a dataframe to multiple dataframes """
    project_cols = d_frame.iloc[row_level-1,:].isin(project_name)\
    [d_frame.iloc[row_level-1,:].isin(project_name)].index.values.tolist()
    row_index = [[index for index,val in enumerate(d_frame.iloc[:,\
                  project_cols[index2]].tolist()) if not pd.isna(val)] \
                  for index2, project in enumerate(project_name)]
    d_frame.drop(d_frame.columns[project_cols],axis = 1, inplace=True)
    return [d_frame.iloc[rows] for rows in row_index]
def make_horizontal_multindex(row_level:int, d_frame:object):
    """ Construct multi-index headers on dataframe """
    first_run = True
    while row_level > 0:
        if first_run: d_frame.columns = d_frame.iloc[0, :].tolist()
        else: d_frame.columns = pd.MultiIndex.from_arrays\
              ([d_frame.columns.tolist(), d_frame.iloc[0, :].tolist()])
        d_frame.drop(0, inplace=True)
        d_frame.reset_index(drop=True, inplace=True)
        row_level = row_level - 1; first_run = False
def make_vertical_multindex(col_level:int, d_frame:object, option = False):
    """ Make vertical multindex structure while ffill operation """
    app = False
    while col_level > 0:
        d_frame.iloc[:, 0] = d_frame.iloc[:, 0].ffill()
        d_frame.set_index(d_frame.iloc[:, 0], append=app, inplace=True)
        d_frame.drop(d_frame.columns[0], axis=1, inplace=True)
        if not option: d_frame.dropna(how='all', inplace=True)
        if option and col_level != 1: d_frame.dropna(how='all', inplace=True)
        col_level = col_level - 1; app = True
def nest(d_frame:object) -> dict:
    """ Make a nested dictionary from stacked dataframe """
    if d_frame.index.nlevels == 1: return d_frame.to_dict()[d_frame.columns[0]]
    return {key: nest(d_frame_group.droplevel(0, axis=0)) \
    for key, d_frame_group in d_frame.groupby(level=0, sort=False)}
def change_outer_structure(end_dict:dict,p_name:list, p_work:list, ws_list:list)->dict:
    """ Change worksheet-projects structure to project- worksheets """
    final_dict = {name: {ws: {} for ws in ws_list} for name in p_name}
    for project in p_name:
        for ws in ws_list:
            if ws not in p_work: final_dict[project][ws] = end_dict[ws]
            if ws in p_work: final_dict[project][ws] = end_dict[ws][project]
    return final_dict
def write_m_files(end_dict:dict, out_path:str):
    """ From nested dictionary prepare .m files in output folder iteratively """

    start_text = "% Produced with Config Generator v1.00\n" + "global CN;\n\n"
    line_break = "%" + "-"*60 + "%\n"

    for project, pro_dict in end_dict.items():

        project_path = os.path.join(out_path, project)
        function_path = os.path.join(project_path, "Functions")
        os.mkdir(project_path); os.mkdir(function_path); once= True

        for worksheet, wsh_dict in pro_dict.items():

            # Write config_system.m file from a ready template
            if once:
                system_text = "% Created with Config Generator 1.00\n"
                system_text += "% Hardware information\n\n"
                system_text += "tempX = rb.settings.getPathMeasurements();\n"
                system_text += "[~,~,tempExt] = fileparts(tempX{1});\n"
                system_text += "if strcmpi(tempExt,'.mf4')\n"
                system_text += "Master='';\nelse\n"
                system_text += "%rb.setSystem({'Measurement_channel','\ETK:1'});\n"
                system_text += "%rb.setSystem({'Measurement_channel','\ETK:2'});\n"
                system_text += "%rb.setSystem({'Measurement_channel','\XCP:1'});\n"
                system_text += "rb.setSystem({'Measurement_channel','\XETK:1'});\n"
                system_text += "Master=rb.getSystem('Measurement_channel');\n"
                system_text += "end"
                system_path = os.path.join(project_path, "config_system.m")
                with open(system_path,'w') as f: f.write(system_text); once=False

            # Write config_signals.m file from Signals worksheet
            if worksheet == "Signals":
                signal_text = "% Produced with Config Generator v1.00\n"
                signal_text += "tempSignals = rb.signals;\n"
                signal_text += "global CN;\n\n"
                for signal_name, signal_dict in wsh_dict.items():
                    signal_eatb = signal_name.split("$")[0]
                    signal_a2l = signal_dict['A2L label']
                    signal_raster = signal_dict["Raster"]
                    signal_text += f"tempSignals.addSignal('{signal_eatb}','{signal_a2l}', Master);\n"
                signal_path = os.path.join(project_path, "config_signals.m")
                with open(signal_path,'w') as f: f.write(signal_text)

            # Write project_name.lab file from Signals worksheet
            if worksheet == "Signals":
                lab_name = file_name.split("_")[1]+ "_"+ file_name.split("_")[2]+ "_"+ project+ ".lab"
                lab_path = os.path.join(project_path, lab_name)
                lab_text = "[RAMCELL]\n\n"
                for eatb_name, eatb_dict in wsh_dict.items():
                    eatb = eatb_name.split("$")[0]
                    a2l  = eatb_dict["A2L label"]
                    raster = eatb_dict["Raster"]
                    lab_text += f"{a2l};{raster} synchronous;\n"
                with open(lab_path,'w') as f: f.write(lab_text)

            # Write config_filters.m from Filters worksheet
            if worksheet == "Filters":
                filter_text = "% Produced with Config Generator v1.00\n"
                for filter_name, filter_dict in wsh_dict.items():
                    filter_type = filter_dict["Type"]
                    filter_para = filter_dict["Parameters"]
                    filter_text+=f"rb.defineFilter('{filter_type}', '{filter_name}', {filter_para});\n"
                filter_path = os.path.join(project_path, "config_filters.m")
                with open(filter_path,'w') as f: f.write(filter_text)

            # Write functions folder content from Functions worksheet
            if worksheet == "Functions":
                for function_name, function_dict in wsh_dict.items():
                    f_name = function_name.split("(")[0]
                    f_path = os.path.join(function_path, f_name + '.m')
                    inputs = function_name.split("(")[1].split(")")[0]
                    output = inputs.split(",")[0]
                    f_text = f"function [result,err] = {f_name}({inputs})\n"
                    f_text += f"result = {output};\nerr = 0;\n\ntry\n"
                    once = True; condition = False
                    for signal_name, f_dict in function_dict.items():
                        sig = signal_name.split("$")[0]
                        if sig == "return": sig = "result"
                        cal = f_dict["Calculation"]
                        con = f_dict["Condition"]
                        if con is not None: condition = True
                        if sig != "None" and cal is not None and con is None:
                            f_text += f"\t{sig} = {cal};\n"
                        if sig == "None" and cal is not None and con is None:
                            f_text += f"{cal}\n"
                        if sig == "None" and cal is not None and con is not None:
                            if once:
                                if con != "else":
                                    f_text += f"\tif {con}\n"
                                    f_text += f"{cal}\n"
                                if con == "else":
                                    f_text += "\telse\n"
                                    f_text += f"{cal}\n"
                                    once = False
                            else: f_text += f"{cal}\n"
                            if condition and once: f_text += "\tend\n"
                        if sig != "None" and cal is not None and con is not None:
                            if once:
                                if con != "else":
                                    f_text += f"\tif {con}\n"
                                    f_text += f"\t\t{sig} = {cal};\n"
                                if con == "else":
                                    f_text += "\telse\n"
                                    f_text += f"\t\t{sig} = {cal};\n"
                                    once = False
                            else: f_text += f"\t\t{sig} = {cal};\n"
                    if condition and not once: f_text += "\tend\n"
                    f_text += "catch me\n\terr = 1;\nend\n\nclear temp*"
                    with open(f_path,'w') as f: f.write(f_text)

            # Write config_Diffs.m from Calculation worksheet
            if worksheet == "Calculation":
                d_text = start_text
                d_text += "Const.WarningNoSignal = 'Missing signals in ';\n"
                d_text += "Const.WarningCrash = 'Something went wrong in ';\n"
                d_text += "Const.InfoTime = 'Calculation time for ';\n"
                d_text += "Const.dt = rbtb.getGrid();\n"
                for block, block_dict in wsh_dict.items():
                    cal_list = [dd["Calculation"]["Calculation"] for dd in block_dict.values()]
                    cal_list = [val for val in cal_list if r"<<" in str(val) and r">>" in str(val)]
                    sig_list = [re.findall(r'<<(.*?)>>', call) for call in cal_list]
                    sig_list = ["<<"+val+">>" for m_list in sig_list for val in m_list]
                    sig_flag = True if len(sig_list)>0 else False
                    d_text += f"try\n\ttempBlockName = string('{block}');\n\ttic;\n\t$USE\n"
                    if sig_flag:
                        sig_txt = "if "
                        for index, c_sig in enumerate(sig_list):
                            if index != len(sig_list)-1: sig_txt += f"isempty({c_sig}) || "
                            else: sig_txt += f"isempty({c_sig})"
                        d_text += f"\t{sig_txt}\n"
                        d_text += "\t\twarning(char(Const.WarningNoSignal + tempBlockName));\n"
                        d_text += "\telse\n"
                    con_memory = "very_different_sentence_with_numbers"; once= False
                    for signal, signal_dict in block_dict.items():
                        end_flag = True if list(block_dict.keys()).index(signal) \
                                == len(list(block_dict.keys())) - 1 else False
                        next_flag = False; next_cond = False
                        if not end_flag:
                            index = list(block_dict.keys()).index(signal)
                            next_S = list(block_dict.keys())[index+1]
                            next_C = block_dict[next_S]["Condition"]["Condition"]
                            next_flag = True if next_C == None else False
                            curr_C = block_dict[signal]["Condition"]["Condition"]
                            if curr_C == "else" and next_C != "else" and not str(next_C).\
                                startswith("elseif") and not next_C is None: next_cond=True

                        sig = signal.split("$")[0]
                        cal = signal_dict["Calculation"]["Calculation"]
                        con = signal_dict["Condition"]["Condition"]
                        if con is None:
                            if sig == "None":
                                if sig_flag: d_text += f"\t\t{cal}\n"
                                else: d_text += f"\t{cal}\n"
                            if sig != "None":
                                if sig_flag: d_text += f"\t\t{sig} = {cal};\n"
                                else: d_text += f"\t{sig} = {cal};\n"
                        if con is not None:
                            if sig == "None":
                                if con != "else" and not con.startswith("elseif"):
                                    if sig_flag and con != con_memory: d_text += f"\t\tif {con}\n"
                                    if not sig_flag and con != con_memory: d_text += f"\tif {con}\n"
                                    if sig_flag: d_text += f"\t\t\t{cal}\n"
                                    else: d_text += f"\t\t{cal}\n"
                                if con == "else":
                                    if sig_flag and not once: d_text += "\t\telse\n"
                                    if not sig_flag and not once: d_text += "\telse\n"
                                    if sig_flag: d_text += f"\t\t\t{cal}\n"
                                    else: d_text += f"\t\t{cal}\n"
                                    once = True
                                if con.startswith("elseif"):
                                    if not sig_flag and con!=con_memory: d_text += f"\t\t{con}\n"
                                    if sig_flag and con!=con_memory: d_text += f"\t\t\t{con}\n"
                                    if sig_flag: d_text += f"\t\t\t{cal}\n"
                                    if not sig_flag: d_text += f"\t\t{cal}\n"
                            if sig != "None":
                                if con != "else" and not con.startswith("elseif"):
                                    if sig_flag and con != con_memory: d_text += f"\t\tif {con}\n"
                                    if not sig_flag and con != con_memory: d_text += f"\tif {con}\n"
                                    if sig_flag: d_text += f"\t\t\t{sig} = {cal};\n"
                                    else: d_text += f"\t\t{sig} = {cal};\n"
                                if con == "else":
                                    if sig_flag and not once: d_text += "\t\t\telse\n"
                                    if not sig_flag and not once: d_text += "\t\telse\n"
                                    if sig_flag: d_text += f"\t\t\t{sig} = {cal};\n"
                                    else: d_text += f"\t\t{sig} = {cal};\n"
                                    once = True
                                if con.startswith("elseif"):
                                    if not sig_flag and con!=con_memory: d_text += f"\t\t{con}\n"
                                    if sig_flag and con!=con_memory: d_text += f"\t\t\t{con}\n"
                                    if sig_flag: d_text += f"\t\t\t{sig} = {cal};\n"
                                    if not sig_flag: d_text += f"\t\t{sig} = {cal};\n"

                            if next_cond and not end_flag and not sig_flag: d_text += "\tend\n"
                            if next_cond and not end_flag and sig_flag: d_text += "\t\tend\n"
                            if next_flag and not end_flag and not sig_flag: d_text += "\tend\n"
                            if next_flag and not end_flag and sig_flag: d_text += "\t\tend\n"
                            if end_flag and sig_flag: d_text += "\t\tend\n"
                            if end_flag and not sig_flag: d_text += "\tend\n"
                            if next_cond or next_flag or end_flag: once = False
                            con_memory = con
                    if sig_flag: d_text += "\tend\n\n"
                    d_text += "\tif toc > Const.MaxBlockTime\n"
                    d_text += "\t\tdisp(char(Const.InfoTime + tempBlockName + ': ' + toc));\n"
                    d_text += "\tend\n\n"
                    d_text += "catch ME\n"
                    d_text += "\twarning(char(Const.WarningCrash + tempBlockName));\n"
                    d_text += "\tdisp(ME.message);\nend\n\n"
                d_text += "clear temp*;\n"
                diff_path = os.path.join(project_path, "config_Diffs.m")
                with open(diff_path,'w') as f: f.write(d_text)

            # Write chapter_name.m files in project folders from Graphs worksheet
            if worksheet == "Graphs":
                for cha, cha_dict in wsh_dict.items(): # chapter
                    cha_text = start_text
                    cha_text += f"rb.addChapter('{cha}');\n"
                    cha_text += f"tempChapter = rb.getChapter('{cha}');\n"
                    cha_text += line_break * 3

                    for sec, sec_dict in cha_dict.items(): # section
                        cha_text += f"tempChapter.addSection('{sec}');\n"
                        cha_text += f"tempSection = tempChapter.getSection('{sec}');\n"
                        cha_text += line_break * 2

                        for chn , chn_dict in sec_dict.items(): # chart name
                            chart_name = chn; first_run = True; num = 2

                            for cht, cht_dict in chn_dict.items(): # chart type
                                chart_type = cht.split("$")[0]
                                if first_run: cha_text+= f"tempSection.addChart('{chart_type}','{chart_name}');\n"
                                if first_run: cha_text+= f"tempChart = tempSection.getChart('{chart_name}');\n"
                                if not first_run: cha_text += f"tempSection.addChild('{chart_type}','{chart_name}_{num}', '{chart_name}');\n"
                                if not first_run: cha_text += f"tempChart = tempSection.getChart('{chart_name}_{num}');\n"
                                if not first_run: num += 1

                                for sig, sig_dict in cht_dict.items(): # signal
                                    signal_name = sig.split("$")[0]
                                    adv_config = sig_dict["Advanced config"]["Advanced config"]

                                    pos_x  = sig_dict["Position"]["X"]
                                    pos_y  = sig_dict["Position"]["Y"]
                                    pos_dx = sig_dict["Position"]["dX"]
                                    pos_dy = sig_dict["Position"]["dY"]

                                    # dis_fun = sig_dict["Display"]["Chart function"]
                                    dis_typ = sig_dict["Display"]["Type"]
                                    dis_tex = sig_dict["Display"]["Text block"]

                                    tri_sig = sig_dict["Trigger"]["Signal"]
                                    tri_thr = sig_dict["Trigger"]["Threshold"]
                                    tri_tim = sig_dict["Trigger"]["Time-shift"]
                                    tri_dur = sig_dict["Trigger"]["Duration"]
                                    tri_che = sig_dict["Trigger"]["Check"]

                                    con_che = sig_dict["Condition"]["Check"]
                                    con_pos = sig_dict["Condition"]["Post-delay"]
                                    con_pre = sig_dict["Condition"]["Pre-delay"]
                                    con_sig = sig_dict["Condition"]["Signal"]
                                    con_thr = sig_dict["Condition"]["Threshold"]
                                    break

                                all_signal = list(cht_dict.keys())
                                sta_signal = [sig for sig in all_signal if sig.startswith("DFES")]
                                exp_signal = [sig for sig in all_signal if sig not in sta_signal]
                                sta_signal = [sig for sig in sta_signal if sig.split("$")[0]!="None"]
                                exp_signal = [sig for sig in exp_signal if sig.split("$")[0]!="None"]
                                sta_text = str([sig.split("$")[0] for sig in sta_signal])[1:-1]
                                exp_text = str([sig.split("$")[0] for sig in exp_signal])[1:-1]

                                if exp_signal: cha_text+= f"tempChart.setSignals({exp_text});\n"
                                if sta_signal: cha_text+= f"tempChart.setStatusSignals({sta_text});\n"

                                if con_che is not None and con_pos is not None and con_pre is not None \
                                                       and con_sig is not None and con_thr is not None:
                                    if con_sig != con_sig.split("\n"):
                                        con_s = con_sig.split("\n")
                                        con_c = con_che.split("\n")
                                        con_t = str(con_thr).split("\n")
                                        con_pr= str(con_pre).split("\n")
                                        con_po= str(con_pos).split("\n")
                                        for i in range(len(con_s)):
                                            if i != len(con_s)-1:
                                                cha_text += f"tempChart.addCondition('{con_s[i][:-1]}', '{con_c[i][:-1]}', {con_t[i][:-1]}, {con_pr[i][:-1]}, {con_po[i][:-1]});\n"
                                            else: cha_text += f"tempChart.addCondition('{con_s[i]}', '{con_c[i]}', {con_t[i]}, {con_pr[i]}, {con_po[i]});\n"
                                    else: cha_text += f"tempChart.addCondition('{con_sig}', '{con_che}', {con_thr}, {con_pre}, {con_pos});\n"


                                if con_che is not None and con_pos is None and con_pre is None \
                                                       and con_sig is not None and con_thr is not None:
                                    if con_sig != con_sig.split("\n"):
                                        con_s = con_sig.split("\n")
                                        con_c = con_che.split("\n")
                                        con_t = str(con_thr).split("\n")
                                        for i in range(len(con_s)):
                                            if i != len(con_s)-1:
                                                cha_text += f"tempChart.addCondition('{con_s[i][:-1]}', '{con_c[i][:-1]}', {con_t[i][:-1]});\n"
                                            else: cha_text += f"tempChart.addCondition('{con_s[i]}', '{con_c[i]}', {con_t[i]});\n"
                                    else: cha_text += f"tempChart.addCondition('{con_sig}', '{con_che}', {con_thr});\n"

                                if tri_sig is not None and tri_che is not None and tri_thr is not None \
                                    and tri_tim is not None and tri_dur is not None and tri_sig != "" \
                                    and tri_che != "" and tri_thr != "" and tri_tim != "" and tri_dur != "":
                                        cha_text += f"tempChart.addTrigger('{tri_sig}', '{tri_che}', {str(tri_thr)}, {str(tri_tim)}, {str(tri_dur)});\n"

                                qty_list = []; legend = []; second_y = []
                                for i, signal in enumerate(exp_signal):
                                    # Qty
                                    if cht_dict[signal]["Signal(s)"]["Qty"] is not None:
                                        qty_list.append(cht_dict[signal]["Signal(s)"]["Qty"])
                                    # Second y axis
                                    if cht_dict[signal]["Display"]["Signal on second Y-Axis"] is not None:
                                        second_y.append(signal.split("$")[0])
                                    # Legend
                                    if i == len(exp_signal)-1: legend.append(signal.split("$")[0])

                                ran_min = ran_max = None

                                for i, signal in enumerate(exp_signal):
                                    sig_min = cht_dict[signal]["Signal(s)"]["Min"]
                                    sig_low = cht_dict[signal]["Signal(s)"]["Low"]
                                    sig_hig = cht_dict[signal]["Signal(s)"]["High"]
                                    sig_max = cht_dict[signal]["Signal(s)"]["Max"]
                                    sig_beh = cht_dict[signal]["Signal(s)"]["Behavior"]
                                    sig_tol = cht_dict[signal]["Signal(s)"]["Tolerance"]
                                    sig_val = [sig_min, sig_low, sig_hig, sig_max]
                                    dis_val = [0 if sig is None else 1 for sig in sig_val]

                                    ran_min = cht_dict[signal]["Display"]["Range min"]
                                    ran_max = cht_dict[signal]["Display"]["Range max"]

                                    if sig_min is None and sig_low is None \
                                    and sig_hig is None and sig_max is None: continue
                                    for idx, val in enumerate(sig_val):
                                        if idx == 0: continue
                                        if val is None: sig_val[idx] = sig_val[idx-1]
                                    for idx, val in reversed(list(enumerate(sig_val))):
                                        if idx == len(sig_val)-1: continue
                                        if val is None: sig_val[idx] = sig_val[idx+1]
                                    sig_txt = str(sig_val).replace("'", "")
                                    dis_txt = str(dis_val).replace("'", "")

                                    if sig_beh is not None and sig_beh != "" and sig_beh!= "None" and sig_tol is None:
                                        if chart_type != "minmax": cha_text += f"tempChart.setSignalThresholds({i+1}, {sig_txt},'{sig_beh}');\n"
                                        if chart_type == "minmax": cha_text += f"tempChart.setSignalThresholds({i+1}, {sig_txt});\n"
                                    if sig_beh is None and sig_tol is not None:
                                        if chart_type != "minmax": cha_text += f"tempChart.setSignalThresholds({i+1}, {sig_txt},'{sig_tol}');\n"
                                        if chart_type == "minmax": cha_text += f"tempChart.setSignalThresholds({i+1}, {sig_txt});\n"
                                    if sig_beh is None and sig_tol is None:
                                        cha_text += f"tempChart.setSignalThresholds({i+1}, {sig_txt});\n"
                                    if 0 in dis_val:
                                        cha_text += f"tempChart.setDisplayThresholds({i+1}, {dis_txt});\n"
                                    if dis_val == [1,1,1,1]:
                                        if len(sig_val) == 1 or len(sig_val) == 2: cha_text += f"tempChart.setDisplayThresholds({i+1}, {dis_txt});\n"


                                    # if ran_min is not None and ran_max is not None and ran_min != ran_max:
                                    #     cha_text += f"tempChart.setRange([], [{ran_min} {ran_max}]);\n"

                                if qty_list: qty_text = str(qty_list)[1:-1].replace("'", "")
                                if qty_list: cha_text += f"tempChart.setQuantity({qty_text});\n"

                                if ran_min is not None and ran_max is not None and ran_min != ran_max:
                                    cha_text += f"tempChart.setRange([], [{ran_min} {ran_max}]);\n"

                                if second_y: secy_txt = str(second_y)[1:-1].replace("'", "")
                                if second_y: cha_text += f"tempChart.moveSignalToSecondYAxis('{secy_txt}');\n"

                                if dis_typ is not None and dis_typ != "":
                                    cha_text += f"tempChart.displayType = '{dis_typ}';\n"

                                if adv_config is not None: cha_text += f"{adv_config}\n"

                                if dis_tex is not None and dis_tex != "":
                                    divided_lines = dis_tex.split("\n")
                                    for line in divided_lines:
                                        if divided_lines.index(line) != len(divided_lines)-1:
                                            line = line[:-1]
                                        cha_text += f"tempChart.addTextBlock{str(line)};\n"

                                if legend and qty_list: cha_text += f"tempChart.setLegend('{legend[0]}');\n"




                                if first_run: cha_text+= f"tempChart.setPosition({pos_x}, {pos_y}, {pos_dx}, {pos_dy});\n"
                                cha_text+= line_break
                                first_run = False



                    cha_text += "clear temp*;\n"
                    cha_path = os.path.join(project_path, cha + ".m")
                    with open(cha_path,'w') as f: f.write(cha_text)


#%% Find Script Environment Path
env_path = os.getcwd()
excel_list = [file for file in os.listdir(env_path) if file.endswith(".xlsm")]

#%% If output folder exists, delete folder and make output folder
out_path = os.path.join(env_path, "Output")
if os.path.isdir(out_path): sh.rmtree(out_path)
os.mkdir(out_path)

#%% Gather Excel file path and file name
file_path = os.path.join(env_path, excel_list[0])
file_name = os.path.splitext(file_path)[0].rsplit('\\', 1)[-1]

#%% Specify log file name and make log path
log_file = f"{file_name}_Log.xlsx"
log_path = os.path.join(out_path, log_file)

#%% Make the log file and record first logs
start_logger(log_path)
record_log(log_path, "Config Generator has started", "INFO", "", "")
record_log(log_path, "Output folder has been constructed", "INFO", "", "")

#%% Specify filters and exceptions for worksheets
filters = ["list", "template"] # Keywords in names as substring to filter
excepts = ['sample_exception'] # Direct sheet names (case sens.) to except

#%% Read workbook and all worksheet names
wb = op.load_workbook(file_path); ws_list = wb.sheetnames

#%% Filter worksheets and make an empty dataframe dictionary
filter_worksheets(ws_list, filters, excepts); ws_dict = {}
record_log(log_path, f"{len(ws_list)} worksheets are filtered", "INFO", "", "")

#%% Hardcoded Solution for Duplicate Chart Type Columns in Graphs
df = pd.read_excel(file_path, sheet_name="Graphs", header= None).iloc[2:,:]
df.reset_index(drop=True, inplace=True)
df.iloc[:,0].ffill(inplace=True)
df.dropna(subset=list(range(1,df.shape[1])), how='all', inplace=True)
df.reset_index(drop=True, inplace=True)
df.iloc[:,1].ffill(inplace=True)
df.dropna(subset=list(range(2,df.shape[1])), how='all', inplace=True)
df.reset_index(drop=True, inplace=True)
all_list = df.iloc[:,3].tolist()
idx_list = [idx for idx,val in enumerate(all_list) if not pd.isna(val)]
for index,val in enumerate(idx_list.copy()):
    if index != len(idx_list)-1:
        next_val = idx_list[index+1]
        if next_val - val > 1:
            ref_val = all_list[val]
            for i in range(val,next_val):
                all_list[i] = ref_val + f"${index}"

#%% Unmerge cells while copying top-left value and read dataframes
for ws_name in ws_list:
    unmerge_worksheet_copy_top_left_value(wb, ws_name)
    ws_dict[ws_name] = pd.DataFrame(wb[ws_name].values)
record_log(log_path, "All worksheets are unmerged ", "INFO", "", "")

#%% Possible comment column labels, drop comment column if exist
com = ["comment", "comments", "Comment", "Comments", "COMMENT", "COMMENTS"]
drop_comment_column(ws_dict, com); end_dict = {}

#%% Gather project names below project labels and project label
pro = ["project", "projects", "Project", "Projects", "PROJECT", "PROJECTS"]
project_name, project_label = gather_project_names(ws_dict, pro)
record_log(log_path, f"{len(project_name)} Project names are gathered", "INFO", "", "")

#%% Gather project contained sheets and make tuple variable
project_work = gather_project_choice_sheets(ws_dict, project_name)
project_tupp = [tuple((project_label, name)) for name in project_name]

#%% Enter main loop for data manipulation and producing nested dictionary
for w_sheet, d_frame in ws_dict.items():

    # Read worksheet object, drop empty rows and columns
    ws = wb[w_sheet]; drop_empty_rows_and_columns(d_frame)

    # Find outline levels of worksheet object and save constant variable
    row_level, col_level = find_outline_levels(ws); row_archive = row_level

    # Hardcoded Area - Recommended Improvements
    if w_sheet == "Config": row_level = 0
    if w_sheet == "Config": d_frame.dropna(axis=1, how='any', inplace=True)
    if w_sheet == "Graphs": col_level = col_level + 2

    record_log(log_path, f"{w_sheet} sheet has {row_level} horizontal to " +\
                         f"{col_level} vertical outline levels", "INFO", "", "")

    # If header row and first row are equal, drop first row
    drop_duplicate_rows(d_frame, row_level)

    # Forward fill, drop NaN rows (compatible to unique Excel structure)
    for index in range(col_level):
        if index != col_level - 1: d_frame.iloc[:,index].ffill(inplace=True)
        if index == 0: d_frame.iloc[:,index].ffill(inplace=True)
        d_frame.dropna(subset=list(range(index+1,d_frame.shape[1])), how='all', inplace=True)
        d_frame.reset_index(drop=True, inplace=True)

    # Duplicate situation hardcoded solution in chart type column
    if w_sheet == "Graphs": d_frame.iloc[2:,3] = all_list

    # If last index column has duplicate values, add extra information (modify)
    identify_duplicate(d_frame, row_level, col_level, w_sheet)

    # If worksheet has choices, gather row indices and disect
    if not w_sheet in project_work: d_frame = [d_frame]
    else: d_frame = disect_dataframe(d_frame, project_name, row_level)

    # Make multi - index rows (horizontal) and columns (vertical)
    for df in d_frame:
        make_horizontal_multindex(row_level, df)
        make_vertical_multindex(col_level, df)

    # Drop fully empty rows and columns (again after headers constructed)
    for df in d_frame:
        df.dropna(axis=0, how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)

    # Stack/ Melt Dataframe Columns (Add upper headers to left headers)
    var_list = ["app" + str(i + 1) for i in range(row_archive)] # temp
    for index, frame in enumerate(d_frame):
        d_frame[index] = frame.melt(var_name=var_list, \
        ignore_index = False).set_index(var_list, append = True)

    # Use nest function to make nested dictionary from multi index dataframe
    end_dict[w_sheet] = {}
    for index, frame in enumerate(d_frame):
        if len(d_frame) == 0: end_dict[w_sheet] = nest(frame) # frame.fillna('') if json doesnt open
        else: end_dict[w_sheet][project_name[index]] = nest(frame) # frame.fillna('') if json doe...

    # Remove temporary keys in no choice worksheets if availabe (For no project sheets)
    for key1, val1 in end_dict.copy().items():
        for key2, val2 in val1.copy().items():
            if len(val1.keys()) == 1 and key2 == project_name[0] and len(project_name)>1:
                end_dict[key1] = val2

    # Check and remove if extra keys are in sheet (Removable if solved - in Config ws)
    for key1, val1 in end_dict.copy().items():
        for key2, val2 in val1.copy().items():
            for key3, val3 in val2.copy().items():
                if key3 == 1 and type(val3) != dict and len(list(val2.keys()))==1:
                    end_dict[key1][key2] = val3

#%% Make final operations on the structure
end_dict = change_outer_structure(end_dict, project_name, project_work, ws_list)

#%% Write final dictionary into a JSON file
with open(os.path.join(out_path,f"{file_name}.json"), "w") as write_file:
    js.dump(end_dict, write_file, allow_nan=True)

#%% Delete extra variables which are not used later
del row_level, col_level, com, d_frame, df, excel_list, excepts, filters, ws_list
del all_list, frame, i, idx_list, index, key1, key2, key3, next_val, pro, ref_val
del project_tupp, row_archive, val, val1, val2, val3, var_list, w_sheet, wb, ws
del ws_dict, ws_name

#%% Write data to .m files on output folder
write_m_files(end_dict, out_path)
record_log(log_path, "Config Generator has finished", "INFO", "", "")
