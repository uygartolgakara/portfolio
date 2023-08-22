""" 
    This script uses an Excel file as an input. The aim of conditional_deleter function 
    is to collect row indices where column x and column y row values are equal. Then
    uses the index list to delete excel rows iteratively while reducing indices.    
"""

import openpyxl

def conditional_deleter(file_path:str, sheet_name:str, col_x:int, col_y:int, offset=1):
    """ in excel file,in row, compare column x and y values. If equal, delete row """
    
    # define workbook variable by reading file at path
    wb = openpyxl.load_workbook(file_path, read_only=False)
    
    # construct specified sheet's object
    ws = wb[sheet_name]
    
    # find indices where column 1 and 2 values are equal
    indices = [i for i in range(ws.min_row+offset, ws.max_row+1)\
               if ws.cell(row=i,column=col_x).value == ws.cell(row=i,column=col_y).value]
        
    # delete rows while reducing index values with order (because row numbers decrease)
    # edit: in similar works, move from end of the excel rows using reversed()
    for i2 in range(len(indices)): ws.delete_rows(indices[i2]-i2)
    
    # save the workbook in specified path
    wb.save(file_path)

""" Example script run scenario: """
# In desktop file, in sheet 1, compare column index 1 and 2 and delete rows
# conditional_deleter(r"C:\Users\Smith\Desktop\Test.xlsx" ,"Sheet1", 1, 2)
 