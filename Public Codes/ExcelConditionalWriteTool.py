""" 
    This script uses an Excel file as an input. The aim of compare_write function 
    is to collect indices of reference column (indices) inside reference sheet, 
    compare indices to another sheet's index values. If index values are same,
    in that row, collects value in a specified tab and writes in a specified tab
    in reference sheet. Also this script supports different header placements.
"""

import openpyxl

def compare_write(file_path:str, ref_sheet:int, ref_tab:int, ref_offset:int, \
                  target_sheet:int, target_ref:int, target_tab:int,          \
                  target_offset:int, final_tab:int):
    """ compares two columns, if equal copy a column to an empty column """
    
    # define workbook variable by reading file at path
    wb = openpyxl.load_workbook(file_path, read_only= False)
    
    # construct specified sheet objects
    ws_ref = wb[ref_sheet]; ws_target = wb[target_sheet]
    
    # collect reference values as a list
    indices = [ws_ref.cell(row=i,column= ref_tab).value \
               for i in range(ws_ref.min_row + ref_offset, ws_ref.max_row + 1)]
        
    for i in range(len(indices)): # repeat reference element number times
        for i2 in range(ws_target.min_row+target_offset, ws_target.max_row+1):
            if indices[i] == ws_target.cell(row=i2,column=target_ref).value:
                
                flag = 1 # check if equal or not and define if condition flag
                # write equal row index values in reference sheet under title
                ws_ref.cell(row=(i+ref_offset+1),column=final_tab).value = \
                ws_target.cell(row=i2,column=target_tab).value
                break
    
    if flag == 1: # if flag is up, write title cell
        # write the title in reference sheet, specified column
        ws_ref.cell(row= ref_offset, column= final_tab).value = \
        ws_target.cell(row= target_offset,column= target_tab).value
        
    wb.save(file_path) # save the final file at same path
                
""" Example script run scenario: """
# In desktop file:
    # read sheet 1 - column 1 as reference data (indices) with one offset (title one row)
    # compare reference data with sheet 3 - column 4 and when reference data is in an index
    # write the value with that index as row and column 5 as data source.
    # Use offset as one again and write the values on sheet 1 - column 9 with title.
# compare_write(r"C:\Users\Smith\Desktop\Test.xlsx","Sheet1", 1, 1, "Sheet2", 3, 4, 1, 10)